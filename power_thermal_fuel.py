"""電力調査統計 火力発電燃料実績の取得・整形。

資源エネルギー庁が公表する統計表 4「火力発電燃料実績」の Excel を年度ごとに取得し、
燃料種×月の 1 行 = 1 レコードへ整形して CSV に保存する。

シート内は燃料種が縦、測定項目（受入量・消費量・発熱量・月末貯蔵量）が横に並ぶ。
単位は燃料種ごとに違い、原典は燃料種の行に単位を書く形で持たせている。
石炭とバイオマスだけは消費量を湿・乾の 2 段で公表するため、燃料種の行の直後に
下段（乾ベース）の消費量だけを持つ行が続く。
"""

import csv
import logging
import re
from pathlib import Path

import openpyxl

from enecho import fetch_file
from power_survey_common import (
    check_sheet,
    number,
    published_as_of,
    resolve_sources,
    squash,
)

logger = logging.getLogger("pipelines")

TABLE_NO = "4"

NAME_HEADING = "燃料種"
UNIT_HEADING_PREFIX = "単位"

# 見出しと列名の対応。2021年4月のシートだけ月末貯蔵量の見出しが「年度末貯蔵量」に
# なっているが、同じシートの単位のセルは月末貯蔵量のままで、値も前月末とは違い
# 翌月へ連続するので月末貯蔵量として読む。
COLUMN_HEADINGS = {
    "受入量": "receipt_quantity",
    "消費量": "consumption_quantity",
    "発熱量": "heat_value",
    "月末貯蔵量": "month_end_stock_quantity",
    "年度末貯蔵量": "month_end_stock_quantity",
}

# 2016年度のシートにだけ「発熱量（加重平均）」の列がある。値が 12 か月すべてで同一
# （月次の実績ではない）で 2017年度以降は列ごと無いため、取り込まずに読み飛ばす。
# 知らない見出しは黙って落とさずに落ちるので、ここに挙げた分だけが読み飛ばされる。
IGNORED_HEADINGS = {"発熱量（加重平均）"}

# 燃料種の名前に付く原典の注記。単位や湿・乾の別は列として持たせるので名前からは外す。
FUEL_NOTE_PATTERN = re.compile(r"（(?:上段|下段)[^）]*）")

# 表の終わりを示す行。以降は注記なので読まない。
REMARK_PREFIX = "備考"

# 単位のセルは「(t)\n（kJ/kg）」の 2 段で、上段が数量、下段が発熱量の単位。
# 原典は上付き文字で 10³m³ と書いており、値だけ取り出すと 103m3 に潰れる。
UNIT_ALIASES = {"103m3": "10^3m3"}

# 単位の英字は半角と全角が混ざる（2024年度の「その他①（単位ｔ報告）」は全角のｔ）。
_FULLWIDTH_ALNUM = str.maketrans(
    {chr(code): chr(code - 0xFEE0) for code in range(0xFF01, 0xFF5F)}
)

VALUE_COLUMNS = [
    "receipt_quantity",
    "consumption_quantity",
    "consumption_dry_quantity",
    "heat_value",
    "month_end_stock_quantity",
]
OUTPUT_COLUMNS = [
    "year_month",
    "fuel_name",
    "quantity_unit",
    "heat_value_unit",
    *VALUE_COLUMNS,
    "published_as_of",
]


def _clean_unit(text: str) -> str:
    """単位のセルの 1 行を単位の文字列にする。原典は括弧で括り全角と半角が混ざる。"""
    unit = squash(text).translate(_FULLWIDTH_ALNUM).strip("()")
    return UNIT_ALIASES.get(unit, unit)


def _parse_units(cell: object) -> tuple[str, str]:
    """単位のセルを (数量の単位, 発熱量の単位) にする。"""
    lines = [line for line in str(cell or "").splitlines() if squash(line)]
    if len(lines) != 2:
        raise RuntimeError(f"単位のセルが 2 段になっていない: {cell!r}")
    return _clean_unit(lines[0]), _clean_unit(lines[1])


def _resolve_columns(rows: list[tuple], sheet_name: str) -> tuple[int, dict[str, int]]:
    """見出し行の位置と、列名から列位置への対応を返す。"""
    for index, row in enumerate(rows):
        if squash(row[0] if row else None) != NAME_HEADING:
            continue
        resolved: dict[str, int] = {}
        for column, value in enumerate(row):
            heading = squash(value)
            if not heading or heading in IGNORED_HEADINGS:
                continue
            if heading.startswith(UNIT_HEADING_PREFIX):
                resolved["_unit"] = column
                continue
            if heading == NAME_HEADING:
                continue
            name = COLUMN_HEADINGS.get(heading)
            if name is None:
                raise RuntimeError(f"{sheet_name}: 未知の見出し '{heading}'（列 {column}）")
            if name in resolved:
                raise RuntimeError(f"{sheet_name}: 見出し '{heading}' が重複している")
            resolved[name] = column
        missing = ({"_unit"} | set(COLUMN_HEADINGS.values())) - set(resolved)
        if missing:
            raise RuntimeError(f"{sheet_name}: 見出しが足りない（{'・'.join(sorted(missing))}）")
        return index, resolved
    raise RuntimeError(f"{sheet_name}: 見出し行が見つからない")


def _cell(row: tuple, index: int) -> object:
    return row[index] if index < len(row) else None


def _parse_sheet(rows: list[tuple], sheet_name: str, year: int, month: int) -> list[tuple]:
    """1 か月のシートを燃料種ごとの行リストへ展開する。"""
    header, columns = _resolve_columns(rows, sheet_name)
    unit_column = columns.pop("_unit")
    as_of = published_as_of(rows[0]) if rows else None
    year_month = f"{year}{month:02d}"

    out: list[dict] = []
    for row in rows[header + 1 :]:
        name = squash(_cell(row, 0))
        if name.startswith(REMARK_PREFIX):
            break
        values = {key: number(_cell(row, index)) for key, index in columns.items()}
        if not name:
            # 燃料種の無い行は石炭・バイオマスの下段（乾ベースの消費量）。
            # 発熱量は 2023年度以前は上段（湿）の行に置かれ、2024年度以降は下段に移った。
            # 消費量と発熱量以外が入っていたら段の対応を読み違えている。
            if all(value is None for value in values.values()):
                continue
            filled = {key for key, value in values.items() if value is not None}
            if not out or not filled <= {"consumption_quantity", "heat_value"}:
                raise RuntimeError(f"{sheet_name}: 読み方の分からない行がある（{values}）")
            if "consumption_quantity" not in filled:
                raise RuntimeError(f"{sheet_name}: 下段に消費量が無い（{values}）")
            if out[-1]["consumption_dry_quantity"] is not None:
                raise RuntimeError(f"{sheet_name}: {out[-1]['fuel_name']} の下段が 2 行ある")
            out[-1]["consumption_dry_quantity"] = values["consumption_quantity"]
            if "heat_value" in filled:
                if out[-1]["heat_value"] is not None:
                    raise RuntimeError(f"{sheet_name}: {out[-1]['fuel_name']} の発熱量が 2 つある")
                out[-1]["heat_value"] = values["heat_value"]
            continue
        quantity_unit, heat_value_unit = _parse_units(_cell(row, unit_column))
        out.append(
            {
                "year_month": year_month,
                "fuel_name": FUEL_NOTE_PATTERN.sub("", name),
                "quantity_unit": quantity_unit,
                "heat_value_unit": heat_value_unit,
                "consumption_dry_quantity": None,
                "published_as_of": as_of,
                **values,
            }
        )

    if not out:
        raise RuntimeError(f"{sheet_name}: 燃料種の行が読めない")
    names = [row["fuel_name"] for row in out]
    if len(set(names)) != len(names):
        raise RuntimeError(f"{sheet_name}: 燃料種が重複している")
    return [tuple(row[column] for column in OUTPUT_COLUMNS) for row in out]


def _parse_workbook(path: Path, fiscal_year: int) -> list[tuple]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: list[tuple] = []
    months = 0
    for sheet_name in workbook.sheetnames:
        year_month = check_sheet(sheet_name, fiscal_year)
        if year_month is None:
            continue
        year, month = year_month
        out.extend(
            _parse_sheet(
                list(workbook[sheet_name].iter_rows(values_only=True)), sheet_name, year, month
            )
        )
        months += 1
    workbook.close()
    if not months:
        raise RuntimeError(f"{fiscal_year}年度のファイルに月次シートが無い")
    return out


def download_and_parse(csv_path: Path, work_dir: Path | None = None) -> int:
    """火力発電燃料実績を取得・整形して CSV に書き出し、行数を返す。"""
    work_dir = Path(work_dir) if work_dir else csv_path.parent
    work_dir.mkdir(parents=True, exist_ok=True)

    all_rows: list[tuple] = []
    for fiscal_year, path in resolve_sources(TABLE_NO):
        xlsx_path = work_dir / f"power_thermal_fuel_{fiscal_year}.xlsx"
        fetch_file(path, xlsx_path)
        rows = _parse_workbook(xlsx_path, fiscal_year)
        logger.info(f"  {fiscal_year}年度: {len(rows)} rows ({path})")
        all_rows.extend(rows)

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(OUTPUT_COLUMNS)
        writer.writerows(all_rows)

    return len(all_rows)

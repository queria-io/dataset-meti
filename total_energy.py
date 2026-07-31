"""総合エネルギー統計 時系列表の取得・整形。

資源エネルギー庁が公表する時系列表（1990年度以降）の Excel を取得し、
統計表×項目×系列×年度の1行=1値へ展開して CSV に保存する。

各シートは同じ形をした複数のブロックの積み重ねになっている:
  - ブロック見出し（右端セルに単位や系列名。例 (PJ) / (構成比)）
  - 「年度」行（年度が横に並ぶ）
  - 項目行（項目名 + 各年度の値）
系列ラベル（単位や構成比・前年度比）は公表どおり保持する。実数と比率の
区別は dbt 側（stg）で系列ラベルから導く。
"""

import csv
import logging
import re
from pathlib import Path

import openpyxl

from enecho import fetch_file, fetch_text

logger = logging.getLogger("pipelines")

RESULTS_PATH = "/statistics/total_energy/results.html"

# 時系列表のファイル名は公表年度と確報/速報で変わる（例 stte_jikeiretu2024fykaku.xlsx）。
# URL を固定すると年度更新で切れるため、統計表一覧ページから解決する。
FILE_PATTERN = re.compile(
    r"/statistics/total_energy/xls/stte/stte_jikeiretu(\d{4})fy([a-z]+)\.xlsx"
)

EDITION_LABELS = {"kaku": "確報", "sokuho": "速報"}

# ブロック見出し（右端セルの "(PJ)" など）
LABEL_PATTERN = re.compile(r"^\((.+)\)$")

OUTPUT_COLUMNS = [
    "table_no",
    "table_name",
    "item_name",
    "item_level",
    "series",
    "fiscal_year",
    "value",
    "source_edition",
]


def resolve_source() -> tuple[str, str]:
    """統計表一覧ページから最新の時系列表 URL と版を解決する。"""
    html = fetch_text(RESULTS_PATH)
    matches = FILE_PATTERN.findall(html)
    if not matches:
        raise RuntimeError("時系列表のリンクが統計表一覧ページに見つからない")
    fiscal_year, kind = max(matches, key=lambda m: (int(m[0]), m[1] == "kaku"))
    path = f"/statistics/total_energy/xls/stte/stte_jikeiretu{fiscal_year}fy{kind}.xlsx"
    edition = f"{fiscal_year}年度{EDITION_LABELS.get(kind, kind)}"
    return path, edition


def _block_label(row: tuple) -> str | None:
    """「年度」行の直前の行から系列ラベルを取り出す。"""
    for value in row:
        if isinstance(value, str):
            matched = LABEL_PATTERN.match(value.strip())
            if matched:
                return matched.group(1)
    return None


def _infer_series(rows: list[tuple], start: int, years: dict[int, int]) -> str:
    """見出しが欠けているブロックの系列を内容から判定する。

    「2.最終消費(エネルギー源別)」の構成比ブロックは原典に見出しが無い。
    構成比は合計行を持たず値が 0〜1 に収まるので、その2点を確認できたときだけ
    構成比と判定し、確認できなければ落とす（推測で埋めない）。
    """
    has_total = False
    within_unit_range = True
    index = start
    while index < len(rows):
        raw_name = rows[index][1] if len(rows[index]) > 1 else None
        if not isinstance(raw_name, str) or not raw_name.strip():
            break
        if raw_name.strip().startswith("※"):
            break
        if raw_name.strip() == "合計":
            has_total = True
        for col in years:
            value = rows[index][col] if col < len(rows[index]) else None
            if isinstance(value, (int, float)) and not 0 <= value <= 1:
                within_unit_range = False
        index += 1
    if not has_total and within_unit_range:
        return "構成比"
    raise RuntimeError(f"系列ラベルが無く内容からも判定できない（行 {start}）")


def _parse_sheet(rows: list[tuple], sheet_name: str, edition: str) -> list[tuple]:
    """1シートを縦持ちの行リストへ展開する。"""
    matched = re.match(r"^(\d+)\.(.+)$", sheet_name)
    if not matched:
        raise RuntimeError(f"想定外のシート名: {sheet_name}")
    table_no = int(matched.group(1))
    # シート名の括弧は全角と半角が混在するため半角へ寄せる。
    table_name = matched.group(2).strip().replace("（", "(").replace("）", ")")

    out: list[tuple] = []
    index = 0
    while index < len(rows):
        row = rows[index]
        header = row[1] if len(row) > 1 else None
        if not (isinstance(header, str) and header.strip() == "年度"):
            index += 1
            continue

        series = None
        # 見出しは直前の行にあるが、空行が挟まる場合に備えて2行遡る。
        for back in (1, 2):
            if index - back >= 0:
                series = _block_label(rows[index - back])
                if series:
                    break

        years = {
            col: int(value)
            for col, value in enumerate(row)
            if isinstance(value, int) and 1900 < value < 2100
        }
        if not years:
            raise RuntimeError(f"{sheet_name}: 年度の並びが読めない（行 {index + 1}）")

        if series is None:
            try:
                series = _infer_series(rows, index + 1, years)
            except RuntimeError as e:
                raise RuntimeError(f"{sheet_name}: {e}") from e

        index += 1
        while index < len(rows):
            item_row = rows[index]
            raw_name = item_row[1] if len(item_row) > 1 else None
            if not isinstance(raw_name, str) or not raw_name.strip():
                break
            # ※ で始まる行は注記でデータではない。
            if raw_name.strip().startswith("※"):
                break
            # 全角空白の字下げが内訳項目を表す。
            item_level = 1 if raw_name.startswith("　") else 0
            item_name = raw_name.strip()
            for col, fiscal_year in years.items():
                value = item_row[col] if col < len(item_row) else None
                # 数値セルのみ採用（"N/A" や空欄は欠測としてスキップ）。
                if not isinstance(value, (int, float)):
                    continue
                out.append(
                    (
                        table_no,
                        table_name,
                        item_name,
                        item_level,
                        series,
                        fiscal_year,
                        value,
                        edition,
                    )
                )
            index += 1
    return out


def download_and_parse(csv_path: Path, work_dir: Path | None = None) -> int:
    """時系列表を取得・整形して CSV に書き出し、行数を返す。"""
    work_dir = Path(work_dir) if work_dir else csv_path.parent
    work_dir.mkdir(parents=True, exist_ok=True)

    path, edition = resolve_source()
    logger.info(f"  時系列表: {edition} ({path})")
    xlsx_path = work_dir / "stte_jikeiretu.xlsx"
    fetch_file(path, xlsx_path)

    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    all_rows: list[tuple] = []
    for sheet_name in workbook.sheetnames:
        # Contents は各シートへの目次で統計値を持たない。
        if sheet_name == "Contents":
            continue
        sheet = workbook[sheet_name]
        all_rows.extend(
            _parse_sheet(list(sheet.iter_rows(values_only=True)), sheet_name, edition)
        )

    if not all_rows:
        raise RuntimeError("時系列表から1行も取り出せなかった")

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(OUTPUT_COLUMNS)
        writer.writerows(all_rows)

    return len(all_rows)

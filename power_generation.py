"""電力調査統計 都道府県別発電実績の取得・整形。

資源エネルギー庁が公表する統計表 2-(2)「都道府県別発電実績」の Excel を
年度ごとに取得し、都道府県×月の 1 行 = 1 レコードへ整形して CSV に保存する。

シート内は都道府県が縦、発電所の種別が横に並ぶ。列の構成は年度で変わる
（蓄電池の列は後から足された）ため、列位置ではなく見出しで対応づける。
"""

import csv
import logging
from pathlib import Path

import openpyxl

from enecho import fetch_file
from power_survey_common import (
    PREFECTURE_CODES,
    check_sheet,
    number,
    published_as_of,
    resolve_sources,
    squash,
)

logger = logging.getLogger("pipelines")

TABLE_NO = "2-2"

# 見出し（大分類, 小分類）と列名の対応。小分類の〔〕は再掲であることを表す原典の記法で、
# 対応づけの前に外す。蓄電池は後から足された列で、古い年度には存在しない。
COLUMN_HEADINGS = {
    ("水力発電所", ""): "hydro_mwh",
    ("火力発電所", ""): "thermal_mwh",
    ("原子力発電所", ""): "nuclear_mwh",
    ("新エネルギー等発電所", "風力"): "wind_mwh",
    ("新エネルギー等発電所", "太陽光"): "solar_mwh",
    ("新エネルギー等発電所", "地熱"): "geothermal_mwh",
    ("新エネルギー等発電所", "バイオマス"): "biomass_mwh",
    ("新エネルギー等発電所", "廃棄物"): "waste_mwh",
    ("新エネルギー等発電所", "蓄電池"): "storage_battery_mwh",
    ("新エネルギー等発電所", "計"): "new_energy_mwh",
    ("その他", ""): "other_mwh",
    ("合計", ""): "total_mwh",
}
OPTIONAL_COLUMNS = {"storage_battery_mwh"}

# 見出しの 3 行目。値を持つ列にだけ入るので、表の右端を決めるのに使う。
UNIT_HEADING = "電力量"
NAME_HEADING = "都道府県"

VALUE_COLUMNS = [
    "hydro_mwh", "thermal_mwh", "nuclear_mwh",
    "wind_mwh", "solar_mwh", "geothermal_mwh",
    "biomass_mwh", "waste_mwh", "storage_battery_mwh", "new_energy_mwh",
    "other_mwh", "total_mwh",
]
OUTPUT_COLUMNS = ["year_month", "pref_code", "pref_name", *VALUE_COLUMNS, "published_as_of"]


def _strip_restatement(heading: str) -> str:
    """再掲を表す括弧を外す。〔バイオマス〕は年度によって括弧の有無が変わる。"""
    return heading.strip("〔〕［］[]")


def _header_rows(rows: list[tuple], sheet_name: str) -> tuple[int, int]:
    """見出し行の位置と都道府県名が入る列を返す。"""
    for index, row in enumerate(rows):
        for column, value in enumerate(row):
            if squash(value) == NAME_HEADING:
                return index, column
    raise RuntimeError(f"{sheet_name}: 見出し行が見つからない")


def _resolve_columns(rows: list[tuple], sheet_name: str) -> dict[str, int]:
    """見出しから列名と列位置の対応を作る。

    大分類は結合セルなので右へ引き延ばす。値を持つ列は 3 行目に単位の見出しが入るので、
    そこで表の右端を切る（切らないと引き延ばした大分類が空列にまで伸びる）。
    """
    header, name_column = _header_rows(rows, sheet_name)
    if len(rows) < header + 3:
        raise RuntimeError(f"{sheet_name}: 見出しが 3 行に足りない")
    top_row, sub_row, unit_row = rows[header], rows[header + 1], rows[header + 2]

    resolved: dict[str, int] = {}
    top = ""
    for column in range(len(unit_row)):
        top = squash(top_row[column] if column < len(top_row) else None) or top
        if squash(unit_row[column]) != UNIT_HEADING:
            continue
        sub = _strip_restatement(squash(sub_row[column] if column < len(sub_row) else None))
        name = COLUMN_HEADINGS.get((top, sub))
        if name is None:
            raise RuntimeError(f"{sheet_name}: 未知の見出し '{top}' / '{sub}'（列 {column}）")
        if name in resolved:
            raise RuntimeError(f"{sheet_name}: 見出し '{top}' / '{sub}' が重複している")
        resolved[name] = column

    missing = set(COLUMN_HEADINGS.values()) - OPTIONAL_COLUMNS - set(resolved)
    if missing:
        raise RuntimeError(f"{sheet_name}: 見出しが足りない（{'・'.join(sorted(missing))}）")
    resolved["_name"] = name_column
    return resolved


def _parse_sheet(rows: list[tuple], sheet_name: str, year: int, month: int) -> list[tuple]:
    """1 か月のシートを都道府県ごとの行リストへ展開する。"""
    columns = _resolve_columns(rows, sheet_name)
    name_column = columns.pop("_name")
    as_of = published_as_of(rows[0]) if rows else None
    year_month = f"{year}{month:02d}"

    out: list[tuple] = []
    for row in rows:
        name = squash(row[name_column] if name_column < len(row) else None)
        # 原典の末尾にある合計行（全国計）は都道府県ではないのでここで落ちる。
        if name not in PREFECTURE_CODES:
            continue
        values = []
        for column in VALUE_COLUMNS:
            index = columns.get(column)
            cell = row[index] if index is not None and index < len(row) else None
            values.append(number(cell))
        out.append((year_month, PREFECTURE_CODES[name], name, *values, as_of))

    if len(out) != len(PREFECTURE_CODES):
        raise RuntimeError(f"{sheet_name}: 都道府県が {len(out)} 件しか読めない")
    return out


def _parse_workbook(path: Path, fiscal_year: int) -> list[tuple]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: list[tuple] = []
    months = 0
    for sheet_name in workbook.sheetnames:
        year_month = check_sheet(sheet_name, fiscal_year)
        if year_month is None:
            continue
        year, month = year_month
        out.extend(
            _parse_sheet(list(workbook[sheet_name].iter_rows(values_only=True)),
                         sheet_name, year, month)
        )
        months += 1
    workbook.close()
    if not months:
        raise RuntimeError(f"{fiscal_year}年度のファイルに月次シートが無い")
    return out


def download_and_parse(csv_path: Path, work_dir: Path | None = None) -> int:
    """都道府県別発電実績を取得・整形して CSV に書き出し、行数を返す。"""
    work_dir = Path(work_dir) if work_dir else csv_path.parent
    work_dir.mkdir(parents=True, exist_ok=True)

    all_rows: list[tuple] = []
    for fiscal_year, path in resolve_sources(TABLE_NO):
        xlsx_path = work_dir / f"power_generation_{fiscal_year}.xlsx"
        fetch_file(path, xlsx_path)
        rows = _parse_workbook(xlsx_path, fiscal_year)
        logger.info(f"  {fiscal_year}年度: {len(rows)} rows ({path})")
        all_rows.extend(rows)

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(OUTPUT_COLUMNS)
        writer.writerows(all_rows)

    return len(all_rows)

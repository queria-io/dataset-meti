"""電力調査統計 都道府県別電力需要実績の取得・整形。

資源エネルギー庁が公表する統計表 3-(2)「都道府県別電力需要実績」の Excel を
年度ごとに取得し、都道府県×月の 1 行 = 1 レコードへ整形して CSV に保存する。

シート内は都道府県が縦、契約区分（特別高圧・高圧・低圧・合計）が横に並ぶ。
低圧だけ特定需要（経過措置料金）と自由料金の内訳を持つ。年度が閉じた後に
「年度計」シートが足されるが、月次の合算で再現できるうえ小売電気事業者数が
「－」になるため取り込まない。
"""

import csv
import logging
from pathlib import Path

import openpyxl

from enecho import fetch_file
from power_survey_common import (
    PREFECTURE_CODES,
    check_sheet,
    number,
    published_as_of,
    resolve_sources,
    squash,
)

logger = logging.getLogger("pipelines")

TABLE_NO = "3-2"

# 値の並びは都道府県名の列を起点とした相対位置で、伝統レイアウト（名称が先頭列）と
# 機械判読用レイアウト（時間軸コード等が前に付き名称は 4 列目）で共通している。
VALUE_OFFSETS = {
    "extra_high_demand_mwh": 1,
    "extra_high_retailers": 2,
    "high_demand_mwh": 3,
    "high_retailers": 4,
    "low_demand_mwh": 5,
    "low_regulated_demand_mwh": 6,
    "low_liberalized_demand_mwh": 7,
    "low_retailers": 8,
    "total_demand_mwh": 9,
    "total_retailers": 10,
}

# 契約区分の見出し（都道府県名の列からの相対位置）。全電圧の見出しだけレイアウトで
# 呼び方が変わる。見出しがずれたら値の並びを信用しない。
HEADING_OFFSETS = {1: ("特別高圧",), 3: ("高圧",), 5: ("低圧",), 9: ("合計", "全電圧計")}

# 小売電気事業者数は整数で、需要量と桁も意味も違う。
COUNT_COLUMNS = {"extra_high_retailers", "high_retailers", "low_retailers", "total_retailers"}

OUTPUT_COLUMNS = ["year_month", "pref_code", "pref_name", *VALUE_OFFSETS, "published_as_of"]


def _name_column(rows: list[tuple], sheet_name: str) -> int:
    """都道府県名が入る列を見つける。レイアウトで先頭列か 4 列目かが変わる。"""
    for row in rows:
        for column, value in enumerate(row):
            if squash(value) in PREFECTURE_CODES:
                return column
    raise RuntimeError(f"{sheet_name}: 都道府県名の列が見つからない")


def _check_headings(rows: list[tuple], sheet_name: str, name_column: int) -> None:
    """契約区分の見出し位置が想定どおりか確かめる。ずれたら黙って読まない。"""
    for offset, expected in HEADING_OFFSETS.items():
        column = name_column + offset
        # 見出しは結合セルで 2〜4 行目に散るため、データが始まる前の行をまとめて見る。
        found = {squash(row[column]) for row in rows[:5] if column < len(row)}
        if not found & set(expected):
            raise RuntimeError(
                f"{sheet_name}: 列 {column} に {'/'.join(expected)} の見出しが無い"
            )


def _parse_sheet(rows: list[tuple], sheet_name: str, year: int, month: int) -> list[tuple]:
    """1 か月のシートを都道府県ごとの行リストへ展開する。"""
    name_column = _name_column(rows, sheet_name)
    _check_headings(rows, sheet_name, name_column)
    as_of = published_as_of(rows[0]) if rows else None
    year_month = f"{year}{month:02d}"

    out: list[tuple] = []
    for row in rows:
        name = squash(row[name_column] if name_column < len(row) else None)
        if name not in PREFECTURE_CODES:
            continue
        values = []
        for column, offset in VALUE_OFFSETS.items():
            index = name_column + offset
            value = number(row[index] if index < len(row) else None)
            values.append(int(value) if value is not None and column in COUNT_COLUMNS else value)
        out.append((year_month, PREFECTURE_CODES[name], name, *values, as_of))

    if len(out) != len(PREFECTURE_CODES):
        raise RuntimeError(f"{sheet_name}: 都道府県が {len(out)} 件しか読めない")
    return out


def _parse_workbook(path: Path, fiscal_year: int) -> list[tuple]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: list[tuple] = []
    months = 0
    for sheet_name in workbook.sheetnames:
        year_month = check_sheet(sheet_name, fiscal_year)
        if year_month is None:
            continue
        year, month = year_month
        out.extend(
            _parse_sheet(list(workbook[sheet_name].iter_rows(values_only=True)),
                         sheet_name, year, month)
        )
        months += 1
    workbook.close()
    if not months:
        raise RuntimeError(f"{fiscal_year}年度のファイルに月次シートが無い")
    return out


def download_and_parse(csv_path: Path, work_dir: Path | None = None) -> int:
    """都道府県別電力需要実績を取得・整形して CSV に書き出し、行数を返す。"""
    work_dir = Path(work_dir) if work_dir else csv_path.parent
    work_dir.mkdir(parents=True, exist_ok=True)

    all_rows: list[tuple] = []
    for fiscal_year, path in resolve_sources(TABLE_NO):
        xlsx_path = work_dir / f"power_demand_{fiscal_year}.xlsx"
        fetch_file(path, xlsx_path)
        rows = _parse_workbook(xlsx_path, fiscal_year)
        logger.info(f"  {fiscal_year}年度: {len(rows)} rows ({path})")
        all_rows.extend(rows)

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(OUTPUT_COLUMNS)
        writer.writerows(all_rows)

    return len(all_rows)

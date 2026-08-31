"""電力調査統計 自家用発電所数、出力の取得・整形。

資源エネルギー庁が公表する統計表 5-(1)「自家用発電所数、出力」の Excel を年度ごとに
取得し、地域×時点の 1 行 = 1 レコードへ整形して CSV に保存する。

電気事業者の発電所を数える 1-(1)・1-(2) と対象が違い、こちらは自家用発電所。公表は
月次ではなく 9 月末・3 月末の年 2 回で、1 つのシートに「自家用発電所数」と「最大出力」の
2 つの表が縦に並ぶ。行は都道府県ではなく 10 地域、列は原動力（水力・汽力・ガスタービン・
内燃力・原子力・風力・太陽光・地熱・燃料電池等）で分かれる。
"""

import csv
import logging
from pathlib import Path

import openpyxl

from enecho import fetch_file
from power_survey_common import check_sheet, number, resolve_sources, squash

logger = logging.getLogger("pipelines")

TABLE_NO = "5-1"

# 行の見出し。原典は 10 地域と全国合計だけを持ち、都道府県の内訳は無い。
# 見出しは全角空白を挟むので squash してから照合する。
AREA_NAMES = (
    "北海道", "東北", "関東", "中部", "北陸", "近畿", "中国", "四国", "九州", "沖縄",
)
TOTAL_ROW_HEADING = "全国合計"
NAME_HEADING = "（地域）"

# 1 シートに縦に並ぶ 2 つの表。表題の付け方は年度で変わる
# （「１．自家用発電所数（…）」と「5-(1)-1.自家用発電所数（…）」）ので、
# 番号を当てにせず表の名前で拾う。
BLOCK_TITLES = {"自家用発電所数": "plants", "最大出力": "capacity_kw"}
NOTE_PREFIX = "（注"

# 小分類の見出しと列名の前半の対応。「計」だけは火力と新エネルギー等で重なるので
# 大分類とあわせて解決する。コージェネレーションは火力の内数（再掲）。
SUB_HEADINGS = {
    "水力": "hydro",
    "汽力": "thermal_steam",
    "ガスタービン": "thermal_gas_turbine",
    "内燃力": "thermal_internal_combustion",
    "火力のうち、コージェネレーションの内数": "thermal_cogeneration",
    "原子力": "nuclear",
    "風力": "wind",
    "太陽光": "solar",
    "地熱": "geothermal",
    "蓄電池": "storage_battery",
    "燃料電池等": "other",
    "合計": "total",
}
GROUP_TOTAL_HEADINGS = {"火力": "thermal", "新エネルギー等": "new_energy"}
TOTAL_SUB_HEADING = "計"

# 蓄電池は 2023 年度に足された列で、それ以前の年度には存在しない。
OPTIONAL_CATEGORIES = {"storage_battery"}

CATEGORIES = (
    "hydro",
    "thermal_steam",
    "thermal_gas_turbine",
    "thermal_internal_combustion",
    "thermal",
    "thermal_cogeneration",
    "nuclear",
    "wind",
    "solar",
    "geothermal",
    "storage_battery",
    "new_energy",
    "other",
    "total",
)
MEASURES = ("plants", "capacity_kw")

VALUE_COLUMNS = [f"{category}_{measure}" for category in CATEGORIES for measure in MEASURES]
OUTPUT_COLUMNS = ["year_month", "area_name", *VALUE_COLUMNS]


def _resolve_columns(rows: list[tuple], header: int, sheet_name: str) -> dict[str, int]:
    """見出し 2 行から列名と列位置の対応を作る。

    大分類は結合セルなので、直近の非空セルを右へ引き延ばして「計」の所属を決める。
    引き延ばした大分類は表の右端を越えて伸びるため、小分類が空の列は無視する。
    """
    top_row = rows[header - 1] if header else ()
    sub_row = rows[header]

    resolved: dict[str, int] = {}
    top = ""
    for column, cell in enumerate(sub_row):
        heading = squash(top_row[column] if column < len(top_row) else None)
        if heading:
            top = heading
        sub = squash(cell)
        if not sub or sub == NAME_HEADING:
            continue
        if sub == TOTAL_SUB_HEADING:
            category = GROUP_TOTAL_HEADINGS.get(top)
            if category is None:
                raise RuntimeError(f"{sheet_name}: 「計」の大分類 '{top}' が未知（列 {column}）")
        else:
            category = SUB_HEADINGS.get(sub)
            if category is None:
                raise RuntimeError(f"{sheet_name}: 未知の見出し '{sub}'（列 {column}）")
        if category in resolved:
            raise RuntimeError(f"{sheet_name}: 見出し '{sub}' が重複している")
        resolved[category] = column

    missing = set(CATEGORIES) - OPTIONAL_CATEGORIES - set(resolved)
    if missing:
        raise RuntimeError(f"{sheet_name}: 見出しが足りない（{'・'.join(sorted(missing))}）")
    return resolved


def _parse_block(rows: list[tuple], title: int, sheet_name: str) -> dict[str, list[float | None]]:
    """表題行から 1 つの表を読み、地域ごとの値を大分類の並び順で返す。"""
    header = next(
        (
            index
            for index in range(title + 1, len(rows))
            if any(squash(cell) == NAME_HEADING for cell in rows[index])
        ),
        None,
    )
    if header is None:
        raise RuntimeError(f"{sheet_name}: 表題 {title + 1} 行目の見出し行が見つからない")
    columns = _resolve_columns(rows, header, sheet_name)

    values: dict[str, list[float | None]] = {}
    for row in rows[header + 1 :]:
        name = squash(row[0] if row else None)
        if name == TOTAL_ROW_HEADING:
            break
        if name not in AREA_NAMES:
            continue
        values[name] = [
            number(row[column] if column is not None and column < len(row) else None)
            for column in (columns.get(category) for category in CATEGORIES)
        ]

    if len(values) != len(AREA_NAMES):
        raise RuntimeError(f"{sheet_name}: 地域が {len(values)} 件しか読めない")
    return values


def _parse_sheet(rows: list[tuple], sheet_name: str, year: int, month: int) -> list[tuple]:
    """1 シートの 2 つの表を地域ごとの 1 行にまとめる。"""
    blocks: dict[str, dict[str, list[float | None]]] = {}
    for index, row in enumerate(rows):
        title = squash(row[0] if row else None)
        if title.startswith(NOTE_PREFIX):
            continue
        for name, measure in BLOCK_TITLES.items():
            if name in title:
                if measure in blocks:
                    raise RuntimeError(f"{sheet_name}: 表 '{name}' が 2 つある")
                blocks[measure] = _parse_block(rows, index, sheet_name)
    missing = set(MEASURES) - set(blocks)
    if missing:
        raise RuntimeError(f"{sheet_name}: 表が足りない（{'・'.join(sorted(missing))}）")

    year_month = f"{year}{month:02d}"
    out: list[tuple] = []
    for area in AREA_NAMES:
        values: list[float | None] = []
        for index, category in enumerate(CATEGORIES):
            for measure in MEASURES:
                value = blocks[measure][area][index]
                values.append(int(value) if measure == "plants" and value is not None else value)
        out.append((year_month, area, *values))
    return out


def _parse_workbook(path: Path, fiscal_year: int) -> list[tuple]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: list[tuple] = []
    sheets = 0
    for sheet_name in workbook.sheetnames:
        year_month = check_sheet(sheet_name, fiscal_year)
        if year_month is None:
            continue
        year, month = year_month
        out.extend(
            _parse_sheet(list(workbook[sheet_name].iter_rows(values_only=True)),
                         sheet_name, year, month)
        )
        sheets += 1
    workbook.close()
    if not sheets:
        raise RuntimeError(f"{fiscal_year}年度のファイルに時点のシートが無い")
    return out


def download_and_parse(csv_path: Path, work_dir: Path | None = None) -> int:
    """自家用発電所数、出力を取得・整形して CSV に書き出し、行数を返す。"""
    work_dir = Path(work_dir) if work_dir else csv_path.parent
    work_dir.mkdir(parents=True, exist_ok=True)

    all_rows: list[tuple] = []
    for fiscal_year, path in resolve_sources(TABLE_NO):
        xlsx_path = work_dir / f"power_plants_captive_{fiscal_year}.xlsx"
        fetch_file(path, xlsx_path)
        rows = _parse_workbook(xlsx_path, fiscal_year)
        logger.info(f"  {fiscal_year}年度: {len(rows)} rows ({path})")
        all_rows.extend(rows)

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(OUTPUT_COLUMNS)
        writer.writerows(all_rows)

    return len(all_rows)

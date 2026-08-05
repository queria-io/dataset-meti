"""電力調査統計 市町村別需要電力量・市町村別逆潮流量の取得・整形。

資源エネルギー庁が公表する統計表 6-(1)「市町村別需要電力量」と 6-(2)「市町村別逆潮流量」の
Excel を年度ごとに取得し、市区町村×月の 1 行 = 1 レコードへ整形して CSV に保存する。

2 つの統計表はシートの作りが同じで、市区町村が縦、測定項目が横に並ぶ。見出しは 1 段で、
その 2 行下に単位の見出し（電力量 / 逆潮流量）が値を持つ列にだけ入る。列の構成は年度で
変わる（逆潮流量の蓄電池は後から足された）ため、列位置ではなく見出しで対応づける。
"""

import csv
import logging
from pathlib import Path

import openpyxl

from enecho import fetch_file
from power_survey_common import (
    PREFECTURE_CODES,
    check_sheet,
    published_as_of,
    resolve_sources,
    squash,
)

logger = logging.getLogger("pipelines")

# 見出し行はこの 2 つを持つ行として探す。行の位置は統計表で違う。
PREFECTURE_HEADING = "都道府県"
MUNICIPALITY_HEADING = "市区町村名"

# 原典の備考「単位未満の場合は「α」と記載」。0 ではないが単位（1,000kWh）に満たない値。
BELOW_UNIT = "α"

DEMAND = {
    "table_no": "6-1",
    "unit_heading": "電力量",
    "columns": {
        "特別高圧／高圧": "extra_high_and_high_demand_mwh",
        "低圧": "low_demand_mwh",
        "合計": "total_demand_mwh",
    },
    "optional": set(),
}

REVERSE_FLOW = {
    "table_no": "6-2",
    "unit_heading": "逆潮流量",
    "columns": {
        "水力": "hydro_mwh",
        "火力": "thermal_mwh",
        "原子力": "nuclear_mwh",
        "風力": "wind_mwh",
        "地熱": "geothermal_mwh",
        "太陽光": "solar_mwh",
        "バイオマス": "biomass_mwh",
        "蓄電池": "storage_battery_mwh",
        "その他": "other_mwh",
        "合計": "total_mwh",
    },
    # 蓄電池は2023年度に足された列で、2022年度の表には無い。
    "optional": {"storage_battery_mwh"},
}

KEY_COLUMNS = ["year_month", "pref_code", "pref_name", "municipality_name"]


def output_columns(table: dict) -> list[str]:
    return [*KEY_COLUMNS, *table["columns"].values(), "published_as_of"]


def _header_row(rows: list[tuple], sheet_name: str) -> tuple[int, int, int]:
    """見出し行の位置と、都道府県名・市区町村名が入る列を返す。"""
    for index, row in enumerate(rows):
        squashed = [squash(value) for value in row]
        if PREFECTURE_HEADING in squashed and MUNICIPALITY_HEADING in squashed:
            return (
                index,
                squashed.index(PREFECTURE_HEADING),
                squashed.index(MUNICIPALITY_HEADING),
            )
    raise RuntimeError(f"{sheet_name}: 見出し行が見つからない")


def _resolve_columns(
    rows: list[tuple], sheet_name: str, table: dict
) -> tuple[dict[str, int], int, int]:
    """見出しから列名と列位置の対応を作る。

    値を持つ列には見出しの 2 行下に単位の見出しが入るので、そこを手掛かりに値の列を選ぶ。
    見出しが増えたり読めなくなったらそこで失敗させる（黙って列を落とさない）。
    """
    header, pref_column, name_column = _header_row(rows, sheet_name)
    if len(rows) <= header + 2:
        raise RuntimeError(f"{sheet_name}: 見出しの 2 行下が無い")
    heading_row, unit_row = rows[header], rows[header + 2]

    resolved: dict[str, int] = {}
    for column, value in enumerate(unit_row):
        if squash(value) != table["unit_heading"]:
            continue
        heading = squash(heading_row[column] if column < len(heading_row) else None)
        name = table["columns"].get(heading)
        if name is None:
            raise RuntimeError(f"{sheet_name}: 未知の見出し '{heading}'（列 {column}）")
        if name in resolved:
            raise RuntimeError(f"{sheet_name}: 見出し '{heading}' が重複している")
        resolved[name] = column

    missing = set(table["columns"].values()) - table["optional"] - set(resolved)
    if missing:
        raise RuntimeError(f"{sheet_name}: 見出しが足りない（{'・'.join(sorted(missing))}）")
    return resolved, pref_column, name_column


def _value(cell: object, sheet_name: str, municipality: str) -> tuple[float | None, bool]:
    """値のセルを数値にする。単位未満（α）と空欄はどちらも欠測として NULL にする。"""
    if isinstance(cell, (int, float)):
        return float(cell), False
    if cell is None:
        return None, False
    if squash(cell) == BELOW_UNIT:
        return None, True
    raise RuntimeError(f"{sheet_name} {municipality}: 読めない値 '{cell}'")


def _parse_sheet(
    rows: list[tuple], sheet_name: str, year: int, month: int, table: dict
) -> tuple[list[tuple], int]:
    """1 か月のシートを市区町村ごとの行リストへ展開する。"""
    columns, pref_column, name_column = _resolve_columns(rows, sheet_name, table)
    as_of = published_as_of(rows[0]) if rows else None
    year_month = f"{year}{month:02d}"

    out: list[tuple] = []
    below_unit = 0
    for row in rows:
        pref = squash(row[pref_column] if pref_column < len(row) else None)
        municipality = squash(row[name_column] if name_column < len(row) else None)
        # 見出し行と末尾の備考行はここで落ちる。原典に全国計・都道府県計の行は無い。
        if pref not in PREFECTURE_CODES or not municipality:
            continue
        values = []
        for name in table["columns"].values():
            column = columns.get(name)
            cell = row[column] if column is not None and column < len(row) else None
            value, is_below_unit = _value(cell, sheet_name, municipality)
            below_unit += is_below_unit
            values.append(value)
        out.append((year_month, PREFECTURE_CODES[pref], pref, municipality, *values, as_of))

    prefectures = {row[2] for row in out}
    if len(prefectures) != len(PREFECTURE_CODES):
        raise RuntimeError(f"{sheet_name}: 都道府県が {len(prefectures)} 件しか読めない")
    return out, below_unit


def _parse_workbook(path: Path, fiscal_year: int, table: dict) -> tuple[list[tuple], int]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: list[tuple] = []
    below_unit = 0
    municipalities: list[tuple[str, str]] | None = None
    for sheet_name in workbook.sheetnames:
        year_month = check_sheet(sheet_name, fiscal_year)
        if year_month is None:
            continue
        year, month = year_month
        rows, sheet_below_unit = _parse_sheet(
            list(workbook[sheet_name].iter_rows(values_only=True)),
            sheet_name, year, month, table,
        )
        # 市区町村の並びは年度内で変わらない。ずれたら行を取りこぼしている。
        found = [(row[1], row[3]) for row in rows]
        if municipalities is None:
            municipalities = found
        elif found != municipalities:
            raise RuntimeError(f"{sheet_name}: 市区町村の並びが他の月と違う（{len(found)} 件）")
        out.extend(rows)
        below_unit += sheet_below_unit
    workbook.close()
    if municipalities is None:
        raise RuntimeError(f"{fiscal_year}年度のファイルに月次シートが無い")
    return out, below_unit


def download_and_parse(table: dict, csv_path: Path, work_dir: Path | None = None) -> int:
    """統計表を取得・整形して CSV に書き出し、行数を返す。"""
    work_dir = Path(work_dir) if work_dir else csv_path.parent
    work_dir.mkdir(parents=True, exist_ok=True)
    table_no = table["table_no"]

    all_rows: list[tuple] = []
    for fiscal_year, path in resolve_sources(table_no):
        xlsx_path = work_dir / f"power_{table_no}_{fiscal_year}.xlsx"
        # 一覧ページに載っていても実体が無い年度がある（6-(2) の2025年度）。
        # 落とした年度は黙って消さずに残す。
        if not fetch_file(path, xlsx_path, missing_ok=True):
            logger.warning(f"  {fiscal_year}年度: 一覧に載っているが取得できない（{path}）")
            continue
        rows, below_unit = _parse_workbook(xlsx_path, fiscal_year, table)
        logger.info(f"  {fiscal_year}年度: {len(rows)} rows / 単位未満 {below_unit} セル ({path})")
        all_rows.extend(rows)

    if not all_rows:
        raise RuntimeError(f"統計表 {table_no} が 1 年度も取得できなかった")

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(output_columns(table))
        writer.writerows(all_rows)

    return len(all_rows)

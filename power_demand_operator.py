"""電力調査統計 電力需要実績（事業者別）の取得・整形。

資源エネルギー庁が公表する統計表 3-(1)「電力需要実績」の Excel を年度ごとに取得し、
事業者×月の 1 行 = 1 レコードへ整形して CSV に保存する。

都道府県別（3-(2)）と違い、行が電気事業者で、1 シートに事業者区分ごとの 2 つの表が
縦に並ぶ。みなし小売電気事業者等（旧一般電気事業者）は経過措置料金・最終保障供給・
離島供給の列を持ち、それ以外の小売電気事業者は自由料金の列しか持たない。
見出しは大分類・中分類・小分類の 3 段で、段数と列数が表によって違うため、
列位置ではなく見出しの組み合わせで対応づける。
"""

import csv
import logging
from pathlib import Path

import openpyxl

from enecho import fetch_file
from power_survey_common import (
    check_sheet,
    number,
    published_as_of,
    resolve_sources,
    squash,
)

logger = logging.getLogger("pipelines")

TABLE_NO = "3-1"

NAME_HEADING = "事業者名"
# 2016年度のシートは見出しが「事業社名」になっている（原典の誤記）。
NAME_HEADINGS = {NAME_HEADING, "事業社名"}
TOTAL_ROW_HEADING = "合計"

# 事業者区分の見出しと、収録する区分コード・区分名の対応。
# シートは「○みなし小売電気事業者等」と「○みなし小売電気事業者以外」の 2 つの表からなる。
CATEGORY_HEADINGS = {
    "○みなし小売電気事業者等": ("deemed_retailer", "みなし小売電気事業者等"),
    "○みなし小売電気事業者以外": ("other_retailer", "みなし小売電気事業者以外"),
}

# 事業区分の見出しと列名の対応。原典は該当すれば ○、しなければ 0 か空欄を置く。
FLAG_HEADINGS = {
    ("事業者名", ""): "_name",
    ("小売電気事業者", ""): "is_retail",
    ("一般送配電事業者", ""): "is_general_transmission_distribution",
    ("送電事業者", ""): "is_transmission",
    ("配電事業者", ""): "is_distribution",
    ("特定送配電事業者", ""): "is_specified_transmission_distribution",
    ("発電事業者", ""): "is_generation",
    ("特定卸供給事業者", ""): "is_specified_wholesale",
}

# 需要量の見出し（大分類, 小分類）と列名の対応。大分類は結合セルで右へ引き延ばし、
# 小分類は見出しの 2 段目・3 段目のどちらに置かれていてもよい。
LIBERALIZED = "その他需要（自由料金）"
REGULATED = "特定需要（経過措置料金）"
VALUE_HEADINGS = {
    (LIBERALIZED, "その他需要計"): "liberalized_demand_mwh",
    (LIBERALIZED, "特別高圧"): "liberalized_extra_high_demand_mwh",
    (LIBERALIZED, "高圧"): "liberalized_high_demand_mwh",
    (LIBERALIZED, "低圧計"): "liberalized_low_demand_mwh",
    (LIBERALIZED, "電灯"): "liberalized_low_lighting_demand_mwh",
    (LIBERALIZED, "電力"): "liberalized_low_power_demand_mwh",
    (REGULATED, "特定需要計"): "regulated_demand_mwh",
    (REGULATED, "電灯"): "regulated_lighting_demand_mwh",
    (REGULATED, "電力"): "regulated_power_demand_mwh",
    ("最終保障供給", ""): "last_resort_supply_mwh",
    ("離島供給", ""): "remote_island_supply_mwh",
    ("合計", ""): "total_demand_mwh",
}

# みなし小売電気事業者以外の表に無い列。経過措置料金・最終保障供給・離島供給は
# 旧一般電気事業者と一般送配電事業者の義務なので、そもそも欄が置かれない。
OTHER_RETAILER_ABSENT = {
    "regulated_demand_mwh",
    "regulated_lighting_demand_mwh",
    "regulated_power_demand_mwh",
    "last_resort_supply_mwh",
    "remote_island_supply_mwh",
    "total_demand_mwh",
}

# 年度によって存在しない列。配電事業者と特定卸供給事業者は 2022 年 4 月施行の
# 電気事業法改正で新設された区分で、それ以前の年度には欄そのものが無い。
OPTIONAL_COLUMNS = {"is_distribution", "is_specified_wholesale"}

FLAG_COLUMNS = [
    "is_retail",
    "is_general_transmission_distribution",
    "is_transmission",
    "is_distribution",
    "is_specified_transmission_distribution",
    "is_generation",
    "is_specified_wholesale",
]
VALUE_COLUMNS = [
    "liberalized_demand_mwh",
    "liberalized_extra_high_demand_mwh",
    "liberalized_high_demand_mwh",
    "liberalized_low_demand_mwh",
    "liberalized_low_lighting_demand_mwh",
    "liberalized_low_power_demand_mwh",
    "regulated_demand_mwh",
    "regulated_lighting_demand_mwh",
    "regulated_power_demand_mwh",
    "last_resort_supply_mwh",
    "remote_island_supply_mwh",
    "total_demand_mwh",
]
OUTPUT_COLUMNS = [
    "year_month",
    "operator_category",
    "operator_category_ja",
    "operator_name",
    *FLAG_COLUMNS,
    *VALUE_COLUMNS,
    "published_as_of",
]

# 事業区分の ○ は年度・行によって U+25CB と U+3007 が混ざる。該当しない欄は 0 か空欄。
YES_MARKS = {"○", "〇"}


def _cell(row: tuple, column: int | None) -> object:
    return row[column] if column is not None and column < len(row) else None


def _block_starts(rows: list[tuple], sheet_name: str) -> list[tuple[int, str, str]]:
    """事業者区分の見出し行を上から順に拾う。"""
    starts = [
        (index, *CATEGORY_HEADINGS[squash(row[0])])
        for index, row in enumerate(rows)
        if row and squash(row[0]) in CATEGORY_HEADINGS
    ]
    found = {code for _, code, _ in starts}
    missing = {code for code, _ in CATEGORY_HEADINGS.values()} - found
    if missing:
        raise RuntimeError(f"{sheet_name}: 事業者区分の表が無い（{'・'.join(sorted(missing))}）")
    return starts


def _header_rows(rows: list[tuple], start: int, sheet_name: str) -> tuple[int, int]:
    """見出し行の位置と、データが始まる行を返す。

    見出しは事業者名の行から数段続き、段数は表によって変わる。事業者名の列が
    埋まった最初の行がデータの先頭なので、そこまでを見出しとして扱う。
    """
    header = next(
        (index for index in range(start, len(rows))
         if rows[index] and squash(rows[index][0]) in NAME_HEADINGS),
        None,
    )
    if header is None:
        raise RuntimeError(f"{sheet_name}: 見出し行が見つからない")
    data = next(
        (index for index in range(header + 1, len(rows))
         if rows[index] and isinstance(rows[index][0], str) and rows[index][0].strip()),
        None,
    )
    if data is None or data == header + 1:
        raise RuntimeError(f"{sheet_name}: 見出しの下にデータが無い")
    return header, data


def _resolve_columns(
    rows: list[tuple], header: int, data: int, sheet_name: str, category: str
) -> dict[str, int]:
    """見出しから列名と列位置の対応を作る。

    大分類は結合セルなので右へ引き延ばし、小分類は見出しの 2 段目以降から拾う。
    小分類の無い列は大分類だけで引く。どちらにも見出しが無い列は表の外側で、
    値を持たないことを後で確かめる。
    """
    top_row = rows[header]
    width = max(len(row) for row in rows[header:data])

    resolved: dict[str, int] = {}
    top = ""
    for column in range(width):
        own_top = squash(_cell(top_row, column))
        if own_top in NAME_HEADINGS:
            own_top = NAME_HEADING
        top = own_top or top
        subs = [squash(_cell(rows[index], column)) for index in range(header + 1, data)]
        sub = "/".join(label for label in subs if label)
        if not sub and not own_top:
            continue
        name = FLAG_HEADINGS.get((top, sub)) or VALUE_HEADINGS.get((top, sub))
        if name is None:
            raise RuntimeError(f"{sheet_name}: 未知の見出し '{top}' / '{sub}'（列 {column}）")
        if name in resolved:
            raise RuntimeError(f"{sheet_name}: 見出し '{top}' / '{sub}' が重複している")
        resolved[name] = column

    expected = set(FLAG_HEADINGS.values()) | set(VALUE_HEADINGS.values())
    if category == "other_retailer":
        expected -= OTHER_RETAILER_ABSENT
    missing = expected - OPTIONAL_COLUMNS - set(resolved)
    if missing:
        raise RuntimeError(f"{sheet_name}: 見出しが足りない（{'・'.join(sorted(missing))}）")
    unexpected = set(resolved) & (OTHER_RETAILER_ABSENT if category == "other_retailer" else set())
    if unexpected:
        raise RuntimeError(
            f"{sheet_name}: {category} に無いはずの列がある（{'・'.join(sorted(unexpected))}）"
        )
    return resolved


def _flag(cell: object) -> bool:
    """事業区分の欄。○ が該当で、0 や空欄は非該当。"""
    return isinstance(cell, str) and cell.strip() in YES_MARKS


def _parse_block(
    rows: list[tuple], start: int, end: int, sheet_name: str, category: str,
    category_ja: str, year_month: str, as_of: str | None,
) -> list[tuple]:
    """事業者区分ごとの表を 1 行 = 1 事業者へ展開する。"""
    header, data = _header_rows(rows, start, sheet_name)
    columns = _resolve_columns(rows, header, data, sheet_name, category)
    name_column = columns.pop("_name")
    mapped = set(columns.values()) | {name_column}

    out: list[tuple] = []
    closed = False
    for row in rows[data:end]:
        name = _cell(row, name_column)
        if not isinstance(name, str) or not name.strip():
            continue
        # 原典は事業者の並びの末尾に合計行を置き、その下に別の表や注記が続く。
        if squash(name) == TOTAL_ROW_HEADING:
            closed = True
            break
        # 見出しの無い列に値があれば読み落としているので気付けるようにする。
        for column, cell in enumerate(row):
            if column not in mapped and isinstance(cell, (int, float)) and cell:
                raise RuntimeError(f"{sheet_name}: 見出しの無い列 {column} に値がある")
        flags = []
        for column in FLAG_COLUMNS:
            index = columns.get(column)
            # その年度に区分そのものが無ければ「該当しない」ではなく不明。
            flags.append(None if index is None else _flag(_cell(row, index)))
        # 原典の「α」（備考どおり 0.5MWh 未満）は数値にできないので欠測にする。
        values = [number(_cell(row, columns.get(column))) for column in VALUE_COLUMNS]
        out.append((year_month, category, category_ja, name.strip(), *flags, *values, as_of))

    # 合計行にたどり着かないまま表が終わるのは、表の切れ目を読み違えている合図。
    if not closed:
        raise RuntimeError(f"{sheet_name}: {category} の合計行が見つからない")
    if not out:
        raise RuntimeError(f"{sheet_name}: {category} の事業者の行が読めない")
    return out


def _parse_sheet(rows: list[tuple], sheet_name: str, year: int, month: int) -> list[tuple]:
    """1 か月のシートを事業者ごとの行リストへ展開する。

    事業者名は月内でも一意にならない（原典に同名が並ぶ月がある）。
    """
    as_of = published_as_of(rows[0]) if rows else None
    year_month = f"{year}{month:02d}"
    blocks = _block_starts(rows, sheet_name)
    # 表の切れ目は次の区分の見出し行。最後の表は注記まで見てよい。
    ends = [start for start, _, _ in blocks[1:]] + [len(rows)]
    out: list[tuple] = []
    for (start, category, category_ja), end in zip(blocks, ends):
        out.extend(
            _parse_block(rows, start, end, sheet_name, category, category_ja,
                         year_month, as_of)
        )
    return out


def _parse_workbook(path: Path, fiscal_year: int) -> list[tuple]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: list[tuple] = []
    months = 0
    for sheet_name in workbook.sheetnames:
        year_month = check_sheet(sheet_name, fiscal_year)
        if year_month is None:
            continue
        year, month = year_month
        out.extend(
            _parse_sheet(list(workbook[sheet_name].iter_rows(values_only=True)),
                         sheet_name, year, month)
        )
        months += 1
    workbook.close()
    if not months:
        raise RuntimeError(f"{fiscal_year}年度のファイルに月次シートが無い")
    return out


def download_and_parse(csv_path: Path, work_dir: Path | None = None) -> int:
    """電力需要実績（事業者別）を取得・整形して CSV に書き出し、行数を返す。"""
    work_dir = Path(work_dir) if work_dir else csv_path.parent
    work_dir.mkdir(parents=True, exist_ok=True)

    all_rows: list[tuple] = []
    for fiscal_year, path in resolve_sources(TABLE_NO):
        xlsx_path = work_dir / f"power_demand_operator_{fiscal_year}.xlsx"
        fetch_file(path, xlsx_path)
        rows = _parse_workbook(xlsx_path, fiscal_year)
        logger.info(f"  {fiscal_year}年度: {len(rows)} rows ({path})")
        all_rows.extend(rows)

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(OUTPUT_COLUMNS)
        writer.writerows(all_rows)

    return len(all_rows)

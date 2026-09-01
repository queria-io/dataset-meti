"""電力調査統計 自家用発電実績の取得・整形。

資源エネルギー庁が公表する統計表 5-(2)「自家用発電実績」の Excel を年度ごとに取得し、
地域×月×原動力の 1 行 = 1 レコードへ整形して CSV に保存する。

1 年度が上期・下期の 2 シートに分かれ、シート内は月ごとの塊が縦に積まれる。塊の中は
原動力（水力・火力・原子力・新エネルギー等・その他）が縦、地域（経済産業局）が横で、
地域ごとに発電電力量・所内及び損失電力量・送電電力量・自家消費電力量が並ぶ。
発電所数を数える 5-(1) と対象は同じ自家用発電所だが、こちらは発電した電力量。
"""

import csv
import logging
import re
from pathlib import Path

import openpyxl

from enecho import fetch_file
from power_survey_common import number, resolve_sources, squash

logger = logging.getLogger("pipelines")

TABLE_NO = "5-2"

# 取り込む最も古い年度。2016年度は同じ表番号でもレイアウトが別で、
# 収録している値も発電電力量ではなく「自家消費（業務用を除く）と特定供給の合計」の
# 地域×月だけ（原典の注1）なので、系列がつながらない。
FIRST_FISCAL_YEAR = 2017

# 月次の塊を持つシート。名前は年度で西暦と元号が混ざる（29上期 / 2018上期）。
HALF_SHEET_PATTERN = re.compile(r"^(?:[HR]?\d{2}|\d{4})(上期|下期)$")
# 値を持たないシート。年度計（29年度 / 2018年度）と参考表。
SKIPPED_SHEET_PATTERN = re.compile(r"^(?:[HR]?\d{2}|\d{4})年度$|^（参考）")
HALF_MONTHS = {"上期": (4, 5, 6, 7, 8, 9), "下期": (10, 11, 12, 1, 2, 3)}

# 地域の見出しは経済産業局の名前。5-(1) など他の統計表と突き合わせられるよう、
# 同じ 10 地域の呼び方へ寄せる。原典の全国合計の列は取り込まない。
AREA_BY_BUREAU = {
    "北海道経済産業局": "北海道",
    "東北経済産業局": "東北",
    "関東経済産業局": "関東",
    "中部経済産業局": "中部",
    "北陸支局": "北陸",
    "近畿経済産業局": "近畿",
    "中国経済産業局": "中国",
    "四国経済産業局": "四国",
    "九州経済産業局": "九州",
    "沖縄総合事務局": "沖縄",
}
NAME_HEADING = "局名"
TOTAL_COLUMN_HEADING = "全国合計"

# 地域 1 つあたりの列と、その見出し。見出し行は 3 段（測定項目・送電先・卸取引の再掲）で、
# 送電電力量だけが送電先ごとに分かれる。読み違いに気付けるよう見出しを照合する。
MEASURES = (
    ("generation_mwh", "発電電力量（千ｋＷｈ）", ""),
    ("station_use_and_loss_mwh", "所内及び損失電力量（千ｋＷｈ）", ""),
    ("transmission_to_utilities_mwh", "電気事業者等への送電電力量（千ｋＷｈ）", "電気事業者"),
    ("transmission_to_specified_supply_mwh", "", "特定供給の相手方"),
    ("transmission_to_other_operators_mwh", "", "その他事業者"),
    ("transmission_total_mwh", "", "合計"),
    ("transmission_via_exchange_mwh", "", ""),
    ("self_consumption_mwh", "自家消費電力量（千ｋＷｈ）", ""),
)
EXCHANGE_HEADING = "合計の内、卸電力取引所を通じた取引"

# 原動力の大分類（原典の 2 列目）と、その下の小分類（4 列目）。
# 火力だけは 3 列目に「原動力別」と「燃料別」の 2 系統が入る。
GROUP_HYDRO = "水力"
GROUP_THERMAL = "火力"
GROUP_NUCLEAR = "原子力"
GROUP_NEW_ENERGY = "新エネルギー等"
GROUP_OTHER = "その他"
GROUP_TOTAL = "合計"
THERMAL_BY_PRIME_MOVER = "原動力別"
THERMAL_BY_FUEL = "燃料別"
GROUP_THERMAL_FUEL = "火力（燃料別）"

# 半角カナは原典の表記。値として持つと絞り込みにくいので全角へ寄せる。
NAME_ALIASES = {
    "ｶﾞｽﾀｰﾋﾞﾝ": "ガスタービン",
    "ｺｰｼﾞｪﾈﾚｰｼｮﾝの内数": "コージェネレーション",
}
# 再掲の行。火力のコージェネレーションと燃料別は原動力別の置き換え（原典の備考2・3）、
# 新エネルギー等のバイオマスと廃棄物は火力の原動力別からの再掲（同 備考5）。
REFERENCE_THERMAL_NAMES = {"コージェネレーション"}
REFERENCE_NEW_ENERGY_NAMES = {"バイオマス", "廃棄物"}

OUTPUT_COLUMNS = [
    "year_month",
    "area_name",
    "power_source_group",
    "power_source_name",
    "is_reference",
    *(name for name, _, _ in MEASURES),
]


def _resolve_areas(header_row: tuple, sheet_name: str) -> list[tuple[int, str]]:
    """局名の行から地域の先頭列を解決する。全国合計の列は取り込まない。"""
    areas: list[tuple[int, str]] = []
    for column, cell in enumerate(header_row):
        heading = squash(cell)
        if not heading or heading in (NAME_HEADING, TOTAL_COLUMN_HEADING):
            continue
        area = AREA_BY_BUREAU.get(heading)
        if area is None:
            raise RuntimeError(f"{sheet_name}: 未知の局名 '{heading}'（列 {column}）")
        areas.append((column, area))
    missing = set(AREA_BY_BUREAU.values()) - {area for _, area in areas}
    if missing:
        raise RuntimeError(f"{sheet_name}: 地域が足りない（{'・'.join(sorted(missing))}）")
    return areas


def _check_measures(rows: list[tuple], header: int, first_column: int, sheet_name: str) -> None:
    """最初の地域の見出し 3 段を照合し、列の並びが想定どおりか確かめる。

    見出しが変わった年度を黙って読み違えると、送電先の内訳が入れ替わっても気付けない。
    """
    def heading(row: int, column: int) -> str:
        cells = rows[row]
        return squash(cells[column]) if column < len(cells) else ""

    for offset, (name, top, sub) in enumerate(MEASURES):
        column = first_column + offset
        if top and heading(header + 1, column) != squash(top):
            raise RuntimeError(f"{sheet_name}: {name} の見出しが '{top}' ではない")
        if sub and heading(header + 2, column) != squash(sub):
            raise RuntimeError(f"{sheet_name}: {name} の見出しが '{sub}' ではない")
    exchange = first_column + [name for name, _, _ in MEASURES].index("transmission_via_exchange_mwh")
    if heading(header + 3, exchange) != squash(EXCHANGE_HEADING):
        raise RuntimeError(f"{sheet_name}: 卸電力取引所の再掲の見出しが想定と違う")


def _resolve_category(
    group: str, sub: str, name: str, sheet_name: str
) -> tuple[str, str, bool]:
    """行の見出し 3 つを (大分類, 小分類, 再掲か) にする。"""
    name = NAME_ALIASES.get(name, name)
    if group == GROUP_THERMAL:
        if sub == THERMAL_BY_FUEL:
            return GROUP_THERMAL_FUEL, name, True
        if sub != THERMAL_BY_PRIME_MOVER:
            raise RuntimeError(f"{sheet_name}: 火力の系統 '{sub}' が未知")
        return GROUP_THERMAL, name, name in REFERENCE_THERMAL_NAMES
    if group in (GROUP_NEW_ENERGY, GROUP_OTHER):
        # 小分類を持つ大分類。名前が空だと次元の値が空のまま出てしまうので落とす。
        if not name:
            raise RuntimeError(f"{sheet_name}: '{group}' の小分類の欄が空")
        if group == GROUP_OTHER:
            return GROUP_OTHER, name, False
        return GROUP_NEW_ENERGY, name, name in REFERENCE_NEW_ENERGY_NAMES
    if group in (GROUP_HYDRO, GROUP_NUCLEAR, GROUP_TOTAL):
        # 小分類を持たない大分類。名前の欄は空なので大分類をそのまま名前にする。
        if name:
            raise RuntimeError(f"{sheet_name}: '{group}' に想定外の小分類 '{name}'")
        return group, group, False
    raise RuntimeError(f"{sheet_name}: 未知の大分類 '{group}'")


def _block_month(rows: list[tuple], start: int, end: int) -> int | None:
    """塊の左端の欄から月を読む。上期合計・下期合計の塊は None。

    原典は「4」「月」の 2 つのセルに分けて書く。数字で始まらない塊は月次ではない。
    """
    for row in rows[start:end]:
        label = row[0] if row else None
        if isinstance(label, (int, float)):
            return int(label)
        text = squash(label)
        if text:
            return int(text) if text.isdigit() else None
    return None


def _parse_sheet(
    rows: list[tuple], sheet_name: str, fiscal_year: int, half: str
) -> list[tuple]:
    """1 シート（上期または下期）の月次の塊をすべて読む。"""
    header = next(
        (index for index, row in enumerate(rows) if any(squash(cell) == NAME_HEADING for cell in row)),
        None,
    )
    if header is None:
        raise RuntimeError(f"{sheet_name}: 局名の見出し行が見つからない")
    areas = _resolve_areas(rows[header], sheet_name)
    _check_measures(rows, header, areas[0][0], sheet_name)

    # 塊の切れ目。大分類が水力の行で始まり、合計の行で終わる。
    starts = [
        index
        for index in range(header + 4, len(rows))
        if squash(rows[index][1] if len(rows[index]) > 1 else None) == GROUP_HYDRO
    ]
    if not starts:
        raise RuntimeError(f"{sheet_name}: 月次の塊が見つからない")

    out: list[tuple] = []
    months: list[int] = []
    for order, start in enumerate(starts):
        end = starts[order + 1] if order + 1 < len(starts) else len(rows)
        month = _block_month(rows, start, end)
        if month is None:
            continue
        if month not in HALF_MONTHS[half]:
            raise RuntimeError(f"{sheet_name}: {half}に {month}月 の塊がある")
        # 同じ月の塊が 2 つあると行が二重になる。月の集合だけを見ると気付けない。
        if month in months:
            raise RuntimeError(f"{sheet_name}: {month}月 の塊が 2 つある")
        months.append(month)
        year = fiscal_year if month >= 4 else fiscal_year + 1
        out.extend(_parse_block(rows, start, end, sheet_name, areas, f"{year}{month:02d}"))

    expected = set(HALF_MONTHS[half])
    if set(months) != expected:
        raise RuntimeError(
            f"{sheet_name}: 月が揃わない（{'・'.join(str(m) for m in sorted(expected - set(months)))} が無い）"
        )
    return out


def _parse_block(
    rows: list[tuple],
    start: int,
    end: int,
    sheet_name: str,
    areas: list[tuple[int, str]],
    year_month: str,
) -> list[tuple]:
    """1 か月の塊を地域×原動力の行にする。"""
    out: list[tuple] = []
    seen: set[tuple[str, str, str]] = set()
    group = ""
    sub = ""
    closed = False
    for row in rows[start:end]:
        current = squash(row[1] if len(row) > 1 else None)
        if current:
            group, sub = current, ""
        if not group:
            continue
        current_sub = squash(row[2] if len(row) > 2 else None)
        if current_sub:
            sub = current_sub
        source_group, source_name, is_reference = _resolve_category(
            group, sub, squash(row[3] if len(row) > 3 else None), sheet_name
        )
        key = (year_month, source_group, source_name)
        if key in seen:
            raise RuntimeError(f"{sheet_name}: {year_month} に '{source_group}/{source_name}' が 2 行ある")
        seen.add(key)
        for column, area in areas:
            values = [
                number(row[column + offset] if column + offset < len(row) else None)
                for offset in range(len(MEASURES))
            ]
            out.append((year_month, area, source_group, source_name, is_reference, *values))
        # 塊は合計の行で終わる。以降は次の塊か注記なので読まない。
        if group == GROUP_TOTAL:
            closed = True
            break

    if not closed:
        raise RuntimeError(f"{sheet_name}: {year_month} の塊に '{GROUP_TOTAL}' の行が無い")
    return out


def _parse_workbook(path: Path, fiscal_year: int) -> list[tuple]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: list[tuple] = []
    halves: set[str] = set()
    for sheet_name in workbook.sheetnames:
        matched = HALF_SHEET_PATTERN.match(sheet_name)
        if matched is None:
            # 読み方の分からないシートは黙って落とさない。月が欠けても気付けなくなる。
            if not SKIPPED_SHEET_PATTERN.match(sheet_name):
                raise RuntimeError(f"{fiscal_year}年度: 想定外のシート '{sheet_name}'")
            continue
        half = matched.group(1)
        halves.add(half)
        out.extend(
            _parse_sheet(
                list(workbook[sheet_name].iter_rows(values_only=True)),
                sheet_name,
                fiscal_year,
                half,
            )
        )
    workbook.close()
    if halves != set(HALF_MONTHS):
        raise RuntimeError(f"{fiscal_year}年度のファイルに上期・下期が揃っていない")
    return out


def download_and_parse(csv_path: Path, work_dir: Path | None = None) -> int:
    """自家用発電実績を取得・整形して CSV に書き出し、行数を返す。"""
    work_dir = Path(work_dir) if work_dir else csv_path.parent
    work_dir.mkdir(parents=True, exist_ok=True)

    all_rows: list[tuple] = []
    for fiscal_year, path in resolve_sources(TABLE_NO):
        if fiscal_year < FIRST_FISCAL_YEAR:
            continue
        xlsx_path = work_dir / f"power_generation_captive_{fiscal_year}.xlsx"
        fetch_file(path, xlsx_path)
        rows = _parse_workbook(xlsx_path, fiscal_year)
        logger.info(f"  {fiscal_year}年度: {len(rows)} rows ({path})")
        all_rows.extend(rows)

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(OUTPUT_COLUMNS)
        writer.writerows(all_rows)

    return len(all_rows)

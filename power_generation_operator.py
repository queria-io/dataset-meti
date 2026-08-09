"""電力調査統計 発電実績（事業者別）の取得・整形。

資源エネルギー庁が公表する統計表 2-(1)「発電実績」の Excel を年度ごとに取得し、
事業者×月の 1 行 = 1 レコードへ整形して CSV に保存する。

都道府県別（2-(2)）と違い、行が電気事業者で、火力発電所は燃料別の内訳を持つ。
事業区分の列は電気事業法の改正で後から増えた（配電事業者・特定卸供給事業者）ため、
列位置ではなく見出しで対応づける。
"""

import csv
import logging
from pathlib import Path

import openpyxl

from enecho import fetch_file
from power_survey_common import (
    check_sheet,
    number,
    published_as_of,
    resolve_sources,
    squash,
)

logger = logging.getLogger("pipelines")

TABLE_NO = "2-1"

NAME_HEADING = "事業者名"
# 2016年度のシートは見出しが「事業社名」になっている（原典の誤記）。
NAME_HEADINGS = {NAME_HEADING, "事業社名"}
TOTAL_ROW_HEADING = "合計"

# 事業区分の見出しと列名の対応。原典は該当すれば ○、しなければ 0 を置く。
# 配電事業者・特定卸供給事業者は 2022 年 4 月施行の電気事業法改正で新設された区分で、
# それ以前の年度には列そのものが無い。
FLAG_HEADINGS = {
    ("事業者名", ""): "_name",
    ("小売電気事業者", ""): "is_retail",
    ("一般送配電事業者", ""): "is_general_transmission_distribution",
    ("送電事業者", ""): "is_transmission",
    ("配電事業者", ""): "is_distribution",
    ("特定送配電事業者", ""): "is_specified_transmission_distribution",
    ("発電事業者", ""): "is_generation",
    ("特定卸供給事業者", ""): "is_specified_wholesale",
}

# 発電電力量の見出し（大分類, 小分類）と列名の対応。小分類の〔〕は再掲を表す
# 原典の記法で、対応づけの前に外す。
VALUE_HEADINGS = {
    ("水力発電所", "一般"): "hydro_conventional_mwh",
    ("水力発電所", "揚水式"): "hydro_pumped_storage_mwh",
    ("水力発電所", "計"): "hydro_mwh",
    ("火力発電所", "石炭"): "thermal_coal_mwh",
    ("火力発電所", "ＬＮＧ"): "thermal_lng_mwh",
    ("火力発電所", "石油"): "thermal_oil_mwh",
    ("火力発電所", "ＬＰＧ"): "thermal_lpg_mwh",
    ("火力発電所", "その他ガス"): "thermal_other_gas_mwh",
    ("火力発電所", "歴青質混合物"): "thermal_bituminous_mwh",
    ("火力発電所", "その他"): "thermal_other_mwh",
    ("火力発電所", "計"): "thermal_mwh",
    ("原子力発電所", ""): "nuclear_mwh",
    ("新エネルギー等発電所", "風力"): "wind_mwh",
    ("新エネルギー等発電所", "太陽光"): "solar_mwh",
    ("新エネルギー等発電所", "地熱"): "geothermal_mwh",
    ("新エネルギー等発電所", "バイオマス"): "biomass_mwh",
    ("新エネルギー等発電所", "廃棄物"): "waste_mwh",
    ("新エネルギー等発電所", "蓄電池"): "storage_battery_mwh",
    ("新エネルギー等発電所", "計"): "new_energy_mwh",
    ("その他", ""): "other_mwh",
    ("計", ""): "total_mwh",
}

# 年度によって存在しない列。無ければ NULL で埋める（0 ではない）。
OPTIONAL_COLUMNS = {
    "is_distribution",
    "is_specified_wholesale",
    "storage_battery_mwh",
}

FLAG_COLUMNS = [
    "is_retail",
    "is_general_transmission_distribution",
    "is_transmission",
    "is_distribution",
    "is_specified_transmission_distribution",
    "is_generation",
    "is_specified_wholesale",
]
VALUE_COLUMNS = [
    "hydro_conventional_mwh", "hydro_pumped_storage_mwh", "hydro_mwh",
    "thermal_coal_mwh", "thermal_lng_mwh", "thermal_oil_mwh", "thermal_lpg_mwh",
    "thermal_other_gas_mwh", "thermal_bituminous_mwh", "thermal_other_mwh", "thermal_mwh",
    "nuclear_mwh",
    "wind_mwh", "solar_mwh", "geothermal_mwh",
    "biomass_mwh", "waste_mwh", "storage_battery_mwh", "new_energy_mwh",
    "other_mwh", "total_mwh",
]
OUTPUT_COLUMNS = [
    "year_month", "operator_name", *FLAG_COLUMNS, *VALUE_COLUMNS, "published_as_of",
]

# 事業区分の ○ は年度・行によって U+25CB と U+3007 が混ざる。該当しない欄は 0。
YES_MARKS = {"○", "〇"}


def _strip_restatement(heading: str) -> str:
    """再掲を表す括弧を外す。〔バイオマス〕は年度によって括弧の有無が変わる。"""
    return heading.strip("〔〕［］[]")


def _header_row(rows: list[tuple], sheet_name: str) -> tuple[int, int]:
    """見出し行の位置と事業者名が入る列を返す。"""
    for index, row in enumerate(rows):
        for column, value in enumerate(row):
            if squash(value) in NAME_HEADINGS:
                return index, column
    raise RuntimeError(f"{sheet_name}: 見出し行が見つからない")


def _resolve_columns(rows: list[tuple], sheet_name: str) -> tuple[int, dict[str, int]]:
    """見出しから列名と列位置の対応を作る。

    見出しは 2 行。大分類は結合セルなので右へ引き延ばし、小分類のある列は
    （大分類, 小分類）で、小分類の無い列は大分類そのもので引く。
    どちらにも見出しが無い列は表の外側なので、値を持たないことを後で確かめる。
    """
    header, name_column = _header_row(rows, sheet_name)
    if len(rows) < header + 2:
        raise RuntimeError(f"{sheet_name}: 見出しが 2 行に足りない")
    top_row, sub_row = rows[header], rows[header + 1]

    resolved: dict[str, int] = {}
    top = ""
    for column in range(max(len(top_row), len(sub_row))):
        own_top = squash(top_row[column] if column < len(top_row) else None)
        if own_top in NAME_HEADINGS:
            own_top = NAME_HEADING
        top = own_top or top
        sub = _strip_restatement(squash(sub_row[column] if column < len(sub_row) else None))
        # 2016年度の原子力発電所だけ小分類の行に大分類と同じ語が置かれている。
        if sub == top:
            sub = ""
        if not sub and not own_top:
            continue
        name = FLAG_HEADINGS.get((top, sub)) or VALUE_HEADINGS.get((top, sub))
        if name is None:
            raise RuntimeError(f"{sheet_name}: 未知の見出し '{top}' / '{sub}'（列 {column}）")
        if name in resolved:
            raise RuntimeError(f"{sheet_name}: 見出し '{top}' / '{sub}' が重複している")
        resolved[name] = column

    expected = set(FLAG_HEADINGS.values()) | set(VALUE_HEADINGS.values())
    missing = expected - OPTIONAL_COLUMNS - set(resolved)
    if missing:
        raise RuntimeError(f"{sheet_name}: 見出しが足りない（{'・'.join(sorted(missing))}）")
    if resolved["_name"] != name_column:
        raise RuntimeError(f"{sheet_name}: 事業者名の列が定まらない")
    return header, resolved


def _flag(cell: object) -> bool:
    """事業区分の欄。○ が該当で、0 や空欄は非該当。"""
    return isinstance(cell, str) and cell.strip() in YES_MARKS


def _parse_sheet(rows: list[tuple], sheet_name: str, year: int, month: int) -> list[tuple]:
    """1 か月のシートを事業者ごとの行リストへ展開する。

    同じ事業者名が 2 行並ぶ月がある（原典どおり）ので、事業者名は月内でも一意にならない。
    """
    header, columns = _resolve_columns(rows, sheet_name)
    name_column = columns.pop("_name")
    as_of = published_as_of(rows[0]) if rows else None
    year_month = f"{year}{month:02d}"
    mapped = set(columns.values()) | {name_column}

    out: list[tuple] = []
    for row in rows[header + 2:]:
        name = row[name_column] if name_column < len(row) else None
        if not isinstance(name, str) or not name.strip():
            continue
        # 原典は事業者の並びの末尾に合計行を置き、その下に備考と参考の別表が続く。
        if squash(name) == TOTAL_ROW_HEADING:
            break
        # 見出しの無い列に値があれば読み落としているので気付けるようにする。
        for column, cell in enumerate(row):
            if column not in mapped and isinstance(cell, (int, float)) and cell:
                raise RuntimeError(f"{sheet_name}: 見出しの無い列 {column} に値がある")
        flags = []
        for column in FLAG_COLUMNS:
            index = columns.get(column)
            # その年度に区分そのものが無ければ「該当しない」ではなく不明。
            flags.append(None if index is None
                         else _flag(row[index] if index < len(row) else None))
        values = []
        for column in VALUE_COLUMNS:
            index = columns.get(column)
            cell = row[index] if index is not None and index < len(row) else None
            # 原典の「α」（0 より大きく 1,000kWh 未満）は数値にできないので欠測にする。
            values.append(number(cell))
        out.append((year_month, name.strip(), *flags, *values, as_of))

    if not out:
        raise RuntimeError(f"{sheet_name}: 事業者の行が読めない")
    return out


def _parse_workbook(path: Path, fiscal_year: int) -> list[tuple]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: list[tuple] = []
    months = 0
    for sheet_name in workbook.sheetnames:
        year_month = check_sheet(sheet_name, fiscal_year)
        if year_month is None:
            continue
        year, month = year_month
        out.extend(
            _parse_sheet(list(workbook[sheet_name].iter_rows(values_only=True)),
                         sheet_name, year, month)
        )
        months += 1
    workbook.close()
    if not months:
        raise RuntimeError(f"{fiscal_year}年度のファイルに月次シートが無い")
    return out


def download_and_parse(csv_path: Path, work_dir: Path | None = None) -> int:
    """発電実績（事業者別）を取得・整形して CSV に書き出し、行数を返す。"""
    work_dir = Path(work_dir) if work_dir else csv_path.parent
    work_dir.mkdir(parents=True, exist_ok=True)

    all_rows: list[tuple] = []
    for fiscal_year, path in resolve_sources(TABLE_NO):
        xlsx_path = work_dir / f"power_generation_operator_{fiscal_year}.xlsx"
        fetch_file(path, xlsx_path)
        rows = _parse_workbook(xlsx_path, fiscal_year)
        logger.info(f"  {fiscal_year}年度: {len(rows)} rows ({path})")
        all_rows.extend(rows)

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(OUTPUT_COLUMNS)
        writer.writerows(all_rows)

    return len(all_rows)

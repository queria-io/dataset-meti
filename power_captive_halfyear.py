"""電力調査統計 自家用発電所等運転半期報集計結果の取得・整形。

資源エネルギー庁が公表する統計表 5「自家用発電所等実績」の Excel を年度ごとに取得し、
都道府県×月×原動力の 1 行 = 1 レコードへ整形して CSV に保存する。

2024年度から統計表 5-(1)「自家用発電所数、出力」と 5-(2)「自家用発電実績」が
この 1 ファイルへ統合された。1 年度が 1 シートで、全国合計と 47 都道府県のブロックが
縦に積まれ、ブロックの中は原動力が縦・年度計と 12 か月が横に並ぶ。10 地域までしか
持たなかった 5-(1) / 5-(2) と違い、こちらは都道府県別の内訳を持つ。

発電所数を持つのは年度計の欄だけで、月次の欄は最大出力と電力量だけ。1 度の読み取りで
月次の 12 期間を都道府県×月×原動力の CSV へ、年度計の発電所数を都道府県×年度×原動力の
CSV へ書き分ける。年度計の最大出力は期末（3 月）の値（原典の注7）で月次と重なるため持たない。
"""

import csv
import logging
import re
from pathlib import Path

import openpyxl

from enecho import fetch_file
from power_survey_common import PREFECTURE_CODES, number, resolve_named_sources, squash

logger = logging.getLogger("pipelines")

TABLE_LABEL = "自家用発電所等実績"
FILENAME_PATTERN = r"jikahatsu\d{4}\.xlsx"
# 統合ファイルでの公表は2024年度から。2023年度以前は 5-(1) / 5-(2) が別々に配布されている。
FIRST_FISCAL_YEAR = 2024

SHEET_PATTERN = re.compile(r"^(\d{4})年度$")
# 期間の見出し。原典は全角数字で「【２０２４年４月】」「【２０２４年度合計】」と書く。
_FULLWIDTH_DIGITS = str.maketrans("０１２３４５６７８９", "0123456789")
MONTH_PERIOD_PATTERN = re.compile(r"^【(\d{4})年(\d{1,2})月】$")
ANNUAL_PERIOD_PATTERN = re.compile(r"^【(\d{4})年度合計】$")

# 電力量の単位。原典の注8 に「電力量の単位は、ｋＷｈ」とある。5-(2) の 1,000kWh と違うので、
# 注が変わったことに気付かずに読むと 1,000 倍ずれる。
UNIT_NOTE = "電力量の単位は、ｋＷｈ"

TOTAL_AREA_NAME = "全国合計"
NAME_HEADING = "原動力別"
PLANTS_HEADING = "発電所数"

# 期間 1 つあたりの列と、その見出し。年度計だけは先頭に発電所数の列が挟まる。
MEASURES = (
    ("capacity_kw", "(A)最大出力（ｋＷ）"),
    ("generation_kwh", "(B)発電電力量"),
    ("station_use_and_loss_kwh", "(C)所内及び損失電力量"),
    ("transmission_total_kwh", "(D)電気事業者等への送電電力量"),
    ("transmission_to_utilities_kwh", "(D)-1 電気事業者"),
    ("transmission_to_specified_supply_kwh", "(D)-2 特定供給の相手方"),
    ("transmission_to_other_operators_kwh", "(D)-3 その他事業者"),
    ("transmission_via_exchange_kwh", "(D)-4 卸電力取引所を通じた取引（内数）"),
    ("self_consumption_kwh", "(E)自家消費電力量"),
)

# ブロックの行。原典の並びどおりに持ち、見出しが 1 つでも変わったら失敗させる。
# (2)火力 は (2)-1［火力］原動力計 と同じ値の重複行なので、同値を確かめたうえで落とす。
DUPLICATE_LABEL = "(2)火力"
PRIME_MOVER_TOTAL_LABEL = "(2)-1［火力］原動力計"
ROW_SPEC = (
    ("(1)水力", "水力", "水力", False),
    (DUPLICATE_LABEL, None, None, False),
    (PRIME_MOVER_TOTAL_LABEL, "火力", "計", False),
    ("(a)汽力", "火力", "汽力", False),
    ("(b)ガスタービン", "火力", "ガスタービン", False),
    ("(c)内燃力", "火力", "内燃力", False),
    ("(d)コジェネレーション（内数）", "火力", "コージェネレーション", True),
    ("(2)-2［火力］燃料計", "火力（燃料別）", "計", True),
    ("(a)石炭", "火力（燃料別）", "石炭", True),
    ("(b)石油", "火力（燃料別）", "石油", True),
    ("(c)液化石油ガス", "火力（燃料別）", "液化石油ガス", True),
    ("(d)天然ガス", "火力（燃料別）", "天然ガス", True),
    ("(e)その他ガス", "火力（燃料別）", "その他ガス", True),
    ("(f)瀝青質混合物", "火力（燃料別）", "瀝青質混合物", True),
    ("(g)バイオマス", "火力（燃料別）", "バイオマス", True),
    ("(h)廃棄物", "火力（燃料別）", "廃棄物", True),
    ("(i)その他", "火力（燃料別）", "その他", True),
    ("(3)原子力", "原子力", "原子力", False),
    ("(4)新エネルギー", "新エネルギー", "計", False),
    ("(a)風力", "新エネルギー", "風力", False),
    ("(b)太陽光", "新エネルギー", "太陽光", False),
    ("(c)地熱", "新エネルギー", "地熱", False),
    ("(d)バイオマス（再掲）", "新エネルギー", "バイオマス", True),
    ("(6)燃料電池等", "燃料電池等", "燃料電池等", False),
    ("(7)蓄電池", "蓄電池", "蓄電池", False),
    ("合計", "合計", "合計", False),
)

# 全国合計とブロックの和を突き合わせるときの許容差。実測の最大は 1e-6 kWh。
TOTAL_ABS_TOLERANCE = 1e-3

PLANTS_OUTPUT_COLUMNS = [
    "fiscal_year",
    "pref_code",
    "pref_name",
    "power_source_group",
    "power_source_name",
    "is_reference",
    "plant_count",
]

OUTPUT_COLUMNS = [
    "year_month",
    "pref_code",
    "pref_name",
    "power_source_group",
    "power_source_name",
    "is_reference",
    *(name for name, _ in MEASURES),
]


def _cell(row: tuple, column: int) -> object:
    return row[column] if column < len(row) else None


def _check_unit(rows: list[tuple], sheet_name: str) -> None:
    """電力量の単位の注記を確かめる。単位が変わると値が 1,000 倍ずれる。"""
    if not any(UNIT_NOTE in value for row in rows[:20] for value in row if isinstance(value, str)):
        raise RuntimeError(f"{sheet_name}: 電力量の単位の注記（{UNIT_NOTE}）が見つからない")


def _resolve_periods(
    rows: list[tuple], block: int, fiscal_year: int, sheet_name: str
) -> tuple[list[tuple[str, int]], int]:
    """ブロックの見出し 2 行から、月次の期間と測定項目の先頭列、発電所数の列を解決する。

    年度計の期間は発電所数の列が 1 つ多く、月次より列数が違う。列位置を数えずに
    見出しから解決し、想定と違う並びなら失敗させる。
    """
    header_row = rows[block]
    measure_row = rows[block + 1]
    if squash(_cell(measure_row, 0)) != NAME_HEADING:
        raise RuntimeError(f"{sheet_name}: {block + 2} 行目が '{NAME_HEADING}' の見出しではない")

    marks = [
        (column, squash(value).translate(_FULLWIDTH_DIGITS))
        for column, value in enumerate(header_row)
        if column >= 1 and isinstance(value, str) and "【" in value
    ]
    if not marks:
        raise RuntimeError(f"{sheet_name}: {block + 1} 行目に期間の見出しが無い")

    periods: list[tuple[str, int]] = []
    annual: list[int] = []
    for order, (start, label) in enumerate(marks):
        end = marks[order + 1][0] if order + 1 < len(marks) else len(measure_row)
        headings = [squash(_cell(measure_row, column)) for column in range(start, end)]
        has_plants = bool(headings) and headings[0] == squash(PLANTS_HEADING)
        base = start + 1 if has_plants else start
        for offset, (_, heading) in enumerate(MEASURES):
            if squash(_cell(measure_row, base + offset)) != squash(heading):
                raise RuntimeError(f"{sheet_name}: {label} の列 {base + offset} が '{heading}' ではない")
        extra = [h for h in headings[base - start + len(MEASURES) :] if h]
        if extra:
            raise RuntimeError(f"{sheet_name}: {label} に想定外の列（{'・'.join(extra)}）がある")

        if ANNUAL_PERIOD_PATTERN.match(label):
            # 発電所数を持つのは年度計だけ。最大出力は期末の値で 3 月と重なり、電力量は
            # 12 か月の和なので、この期間から採るのは発電所数だけにする。
            if not has_plants:
                raise RuntimeError(f"{sheet_name}: {label} に '{PLANTS_HEADING}' の列が無い")
            annual.append(start)
            continue
        matched = MONTH_PERIOD_PATTERN.match(label)
        if matched is None:
            raise RuntimeError(f"{sheet_name}: 期間の見出し '{label}' が読めない")
        if has_plants:
            raise RuntimeError(f"{sheet_name}: {label} に '{PLANTS_HEADING}' の列がある")
        year, month = int(matched.group(1)), int(matched.group(2))
        if (year if month >= 4 else year - 1) != fiscal_year:
            raise RuntimeError(f"{sheet_name}: {fiscal_year}年度のファイルに {label} が入っている")
        periods.append((f"{year}{month:02d}", base))

    expected = {
        f"{fiscal_year + (0 if month >= 4 else 1)}{month:02d}"
        for month in (4, 5, 6, 7, 8, 9, 10, 11, 12, 1, 2, 3)
    }
    if {year_month for year_month, _ in periods} != expected:
        raise RuntimeError(f"{sheet_name}: 12 か月が揃っていない（{len(periods)} 期間）")
    if len(annual) != 1:
        raise RuntimeError(f"{sheet_name}: 年度計の期間が {len(annual)} 個ある")
    return periods, annual[0]


def _resolve_blocks(rows: list[tuple], sheet_name: str) -> list[tuple[int, str]]:
    """全国合計と都道府県のブロックの先頭行を解決する。"""
    blocks = [
        (index, squash(_cell(row, 0)))
        for index, row in enumerate(rows)
        if isinstance(_cell(row, 1), str) and "【" in _cell(row, 1)
    ]
    if not blocks or blocks[0][1] != TOTAL_AREA_NAME:
        raise RuntimeError(f"{sheet_name}: 最初のブロックが '{TOTAL_AREA_NAME}' ではない")
    names = [name for _, name in blocks[1:]]
    unknown = [name for name in names if name not in PREFECTURE_CODES]
    if unknown:
        raise RuntimeError(f"{sheet_name}: 未知の都道府県（{'・'.join(unknown)}）")
    missing = set(PREFECTURE_CODES) - set(names)
    if missing:
        raise RuntimeError(f"{sheet_name}: 都道府県が足りない（{'・'.join(sorted(missing))}）")
    return blocks


def _check_labels(rows: list[tuple], block: int, area: str, sheet_name: str) -> None:
    """ブロックの行の見出しを原典の並びと突き合わせる。"""
    for offset, (label, *_) in enumerate(ROW_SPEC):
        actual = squash(_cell(rows[block + 2 + offset], 0))
        if actual != label:
            raise RuntimeError(
                f"{sheet_name}: {area} の {block + 3 + offset} 行目が '{label}' ではなく '{actual}'"
            )


def _check_duplicate_row(rows: list[tuple], block: int, area: str, sheet_name: str) -> None:
    """落とす (2)火力 の行が (2)-1［火力］原動力計 と同値であることを確かめる。"""
    labels = [label for label, *_ in ROW_SPEC]
    duplicate = rows[block + 2 + labels.index(DUPLICATE_LABEL)]
    kept = rows[block + 2 + labels.index(PRIME_MOVER_TOTAL_LABEL)]
    for column in range(1, len(kept)):
        left, right = number(_cell(duplicate, column)), number(_cell(kept, column))
        if left is None and right is None:
            continue
        if left is None or right is None or abs(left - right) > TOTAL_ABS_TOLERANCE:
            raise RuntimeError(
                f"{sheet_name}: {area} の '{DUPLICATE_LABEL}' が"
                f" '{PRIME_MOVER_TOTAL_LABEL}' と違う（列 {column}）"
            )


def _check_annual_capacity(
    rows: list[tuple],
    block: int,
    area: str,
    plants_column: int,
    periods: list[tuple[str, int]],
    fiscal_year: int,
    sheet_name: str,
) -> None:
    """年度計の最大出力が期末（3 月）の値と一致することを確かめる。

    一致するので年度計の最大出力は収録していない（原典の注7）。原典が年度計に別の値を
    書き始めたら持たないという判断ごと変わるので、黙って通さずここで失敗させる。
    """
    march = f"{fiscal_year + 1}03"
    base = next(column for year_month, column in periods if year_month == march)
    for offset in range(len(ROW_SPEC)):
        row = rows[block + 2 + offset]
        annual = number(_cell(row, plants_column + 1))
        monthly = number(_cell(row, base))
        if annual is None and monthly is None:
            continue
        if annual is None or monthly is None or abs(annual - monthly) > TOTAL_ABS_TOLERANCE:
            raise RuntimeError(
                f"{sheet_name}: {area} の '{ROW_SPEC[offset][0]}' で年度計の最大出力"
                f" {annual} が {march} の {monthly} と違う"
            )


def _check_national_total(
    rows: list[tuple], blocks: list[tuple[int, str]], sheet_name: str
) -> None:
    """全国合計のブロックが 47 都道府県の和と一致することを確かめる。

    行がずれて読めていれば都道府県の和が合わなくなる。年度計の列（発電所数を含む）も
    突き合わせるので、収録しない期間の読み違いにも気付ける。ブロックの測定項目の見出しは
    全国合計のブロックでしか照合していないので、都道府県の列ずれを捉える網はこれだけ。

    数値でないセル（原典が「－」を書き始めた場合など）があると和を取れず、その組は
    突き合わせられない。網がどれだけ張れたかを数えてログに残し、1 組も突き合わせられなければ
    網が消えているので失敗させる。
    """
    national = blocks[0][0]
    width = len(rows[national + 2])
    compared = skipped = 0
    for offset in range(len(ROW_SPEC)):
        for column in range(1, width):
            total = number(_cell(rows[national + 2 + offset], column))
            parts = [number(_cell(rows[block + 2 + offset], column)) for block, _ in blocks[1:]]
            if total is None or any(part is None for part in parts):
                skipped += 1
                continue
            compared += 1
            summed = sum(parts)
            if abs(total - summed) > max(TOTAL_ABS_TOLERANCE, abs(total) * 1e-9):
                raise RuntimeError(
                    f"{sheet_name}: {ROW_SPEC[offset][0]} の列 {column} で"
                    f" 全国合計 {total} が都道府県の和 {summed} と合わない"
                )
    if not compared:
        raise RuntimeError(f"{sheet_name}: 全国合計と都道府県の和を 1 組も突き合わせられない")
    logger.info(f"  {sheet_name}: 全国合計と突合 {compared} 組（数値でなく飛ばした組 {skipped}）")


def _parse_sheet(
    rows: list[tuple], sheet_name: str, fiscal_year: int
) -> tuple[list[tuple], list[tuple]]:
    _check_unit(rows, sheet_name)
    blocks = _resolve_blocks(rows, sheet_name)
    periods, plants_column = _resolve_periods(rows, blocks[0][0], fiscal_year, sheet_name)
    for block, area in blocks:
        _check_labels(rows, block, area, sheet_name)
        _check_duplicate_row(rows, block, area, sheet_name)
        _check_annual_capacity(
            rows, block, area, plants_column, periods, fiscal_year, sheet_name
        )
    _check_national_total(rows, blocks, sheet_name)

    out: list[tuple] = []
    plants: list[tuple] = []
    for block, area in blocks[1:]:
        for offset, (_, group, name, is_reference) in enumerate(ROW_SPEC):
            if group is None:
                continue
            row = rows[block + 2 + offset]
            key = (PREFECTURE_CODES[area], area, group, name, is_reference)
            count = number(_cell(row, plants_column))
            plants.append((fiscal_year, *key, None if count is None else int(count)))
            for year_month, base in periods:
                values = [number(_cell(row, base + index)) for index in range(len(MEASURES))]
                out.append((year_month, *key, *values))
    return out, plants


def _parse_workbook(path: Path, fiscal_year: int) -> tuple[list[tuple], list[tuple]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: list[tuple] = []
    plants: list[tuple] = []
    sheets = 0
    for sheet_name in workbook.sheetnames:
        matched = SHEET_PATTERN.match(sheet_name)
        # 読み方の分からないシートは黙って落とさない。年度が欠けても気付けなくなる。
        if matched is None:
            raise RuntimeError(f"{fiscal_year}年度: 想定外のシート '{sheet_name}'")
        if int(matched.group(1)) != fiscal_year:
            raise RuntimeError(f"{fiscal_year}年度のファイルに '{sheet_name}' が入っている")
        rows = list(workbook[sheet_name].iter_rows(values_only=True))
        sheet_rows, sheet_plants = _parse_sheet(rows, sheet_name, fiscal_year)
        out.extend(sheet_rows)
        plants.extend(sheet_plants)
        sheets += 1
    workbook.close()
    if not sheets:
        raise RuntimeError(f"{fiscal_year}年度のファイルに年度のシートが無い")
    return out, plants


def _write_csv(csv_path: Path, columns: list[str], rows: list[tuple]) -> int:
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(columns)
        writer.writerows(rows)
    return len(rows)


def download_and_parse(
    csv_path: Path, plants_csv_path: Path, work_dir: Path | None = None
) -> tuple[int, int]:
    """自家用発電所等実績を取得・整形して 2 つの CSV に書き出し、それぞれの行数を返す。

    月次の実績（csv_path）と年度計の発電所数（plants_csv_path）は同じ Excel から採るので、
    1 度の取得で両方を書く。
    """
    work_dir = Path(work_dir) if work_dir else csv_path.parent
    work_dir.mkdir(parents=True, exist_ok=True)

    all_rows: list[tuple] = []
    all_plants: list[tuple] = []
    for fiscal_year, path in resolve_named_sources(FILENAME_PATTERN, TABLE_LABEL):
        if fiscal_year < FIRST_FISCAL_YEAR:
            continue
        xlsx_path = work_dir / f"power_captive_halfyear_{fiscal_year}.xlsx"
        fetch_file(path, xlsx_path)
        rows, plants = _parse_workbook(xlsx_path, fiscal_year)
        logger.info(f"  {fiscal_year}年度: {len(rows)} rows / 発電所数 {len(plants)} rows ({path})")
        all_rows.extend(rows)
        all_plants.extend(plants)

    return (
        _write_csv(csv_path, OUTPUT_COLUMNS, all_rows),
        _write_csv(plants_csv_path, PLANTS_OUTPUT_COLUMNS, all_plants),
    )

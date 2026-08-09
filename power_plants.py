"""電力調査統計 都道府県別発電所数、出力の取得・整形。

資源エネルギー庁が公表する統計表 1-(2)「都道府県別発電所数、出力」の Excel を
年度ごとに取得し、都道府県×月の 1 行 = 1 レコードへ整形して CSV に保存する。

シートの作りは 2-(2) 都道府県別発電実績と同じで、都道府県が縦・電源種別が横に並ぶ。
違いは電源種別ごとに測定項目が 2 つ（発電所数・最大出力計）ある点で、
見出しの 3 行目が単位ではなく測定項目名になっている。
"""

import csv
import logging
from pathlib import Path

import openpyxl

from enecho import fetch_file
from power_survey_common import (
    PREFECTURE_CODES,
    check_sheet,
    number,
    published_as_of,
    resolve_sources,
    squash,
)

logger = logging.getLogger("pipelines")

TABLE_NO = "1-2"

# 見出し（大分類, 小分類）と列名の前半の対応。小分類の〔〕は再掲であることを表す
# 原典の記法で、対応づけの前に外す。蓄電池は後から足された列で、古い年度には存在しない。
COLUMN_HEADINGS = {
    ("水力発電所", ""): "hydro",
    ("火力発電所", ""): "thermal",
    ("原子力発電所", ""): "nuclear",
    ("新エネルギー等発電所", "風力"): "wind",
    ("新エネルギー等発電所", "太陽光"): "solar",
    ("新エネルギー等発電所", "地熱"): "geothermal",
    ("新エネルギー等発電所", "バイオマス"): "biomass",
    ("新エネルギー等発電所", "廃棄物"): "waste",
    ("新エネルギー等発電所", "蓄電池"): "storage_battery",
    ("新エネルギー等発電所", "計"): "new_energy",
    ("その他", ""): "other",
    ("合計", ""): "total",
}
OPTIONAL_CATEGORIES = {"storage_battery"}

# 見出しの 3 行目。電源種別ごとに 2 列あり、値を持つ列にだけ入るので表の右端を決めるのに使う。
MEASURE_HEADINGS = {"発電所数": "plants", "最大出力計": "capacity_kw"}
NAME_HEADING = "都道府県"

VALUE_COLUMNS = [
    f"{category}_{suffix}"
    for category in COLUMN_HEADINGS.values()
    for suffix in MEASURE_HEADINGS.values()
]
OPTIONAL_COLUMNS = {
    f"{category}_{suffix}"
    for category in OPTIONAL_CATEGORIES
    for suffix in MEASURE_HEADINGS.values()
}
OUTPUT_COLUMNS = ["year_month", "pref_code", "pref_name", *VALUE_COLUMNS, "published_as_of"]


def _strip_restatement(heading: str) -> str:
    """再掲を表す括弧を外す。〔バイオマス〕は年度によって括弧の有無が変わる。"""
    return heading.strip("〔〕［］[]")


def _header_rows(rows: list[tuple], sheet_name: str) -> tuple[int, int]:
    """見出し行の位置と都道府県名が入る列を返す。"""
    for index, row in enumerate(rows):
        for column, value in enumerate(row):
            if squash(value) == NAME_HEADING:
                return index, column
    raise RuntimeError(f"{sheet_name}: 見出し行が見つからない")


def _resolve_columns(rows: list[tuple], sheet_name: str) -> dict[str, int]:
    """見出しから列名と列位置の対応を作る。

    大分類・小分類は結合セルなので右へ引き延ばす。値を持つ列は 3 行目に測定項目の
    見出しが入るので、そこで表の右端を切る（切らないと引き延ばした大分類が空列にまで伸びる）。
    """
    header, name_column = _header_rows(rows, sheet_name)
    if len(rows) < header + 3:
        raise RuntimeError(f"{sheet_name}: 見出しが 3 行に足りない")
    top_row, sub_row, measure_row = rows[header], rows[header + 1], rows[header + 2]

    resolved: dict[str, int] = {}
    top = ""
    sub = ""
    for column in range(len(measure_row)):
        heading = squash(top_row[column] if column < len(top_row) else None)
        if heading:
            top, sub = heading, ""
        sub = _strip_restatement(
            squash(sub_row[column] if column < len(sub_row) else None)
        ) or sub
        suffix = MEASURE_HEADINGS.get(squash(measure_row[column]))
        if suffix is None:
            continue
        category = COLUMN_HEADINGS.get((top, sub))
        if category is None:
            raise RuntimeError(f"{sheet_name}: 未知の見出し '{top}' / '{sub}'（列 {column}）")
        name = f"{category}_{suffix}"
        if name in resolved:
            raise RuntimeError(f"{sheet_name}: 見出し '{top}' / '{sub}' / '{suffix}' が重複している")
        resolved[name] = column

    missing = set(VALUE_COLUMNS) - OPTIONAL_COLUMNS - set(resolved)
    if missing:
        raise RuntimeError(f"{sheet_name}: 見出しが足りない（{'・'.join(sorted(missing))}）")
    resolved["_name"] = name_column
    return resolved


def _parse_sheet(rows: list[tuple], sheet_name: str, year: int, month: int) -> list[tuple]:
    """1 か月のシートを都道府県ごとの行リストへ展開する。"""
    columns = _resolve_columns(rows, sheet_name)
    name_column = columns.pop("_name")
    as_of = published_as_of(rows[0]) if rows else None
    year_month = f"{year}{month:02d}"

    out: list[tuple] = []
    for row in rows:
        name = squash(row[name_column] if name_column < len(row) else None)
        # 原典の末尾にある合計行（全国計）は都道府県ではないのでここで落ちる。
        if name not in PREFECTURE_CODES:
            continue
        values = []
        for column in VALUE_COLUMNS:
            index = columns.get(column)
            cell = row[index] if index is not None and index < len(row) else None
            values.append(number(cell))
        out.append((year_month, PREFECTURE_CODES[name], name, *values, as_of))

    if len(out) != len(PREFECTURE_CODES):
        raise RuntimeError(f"{sheet_name}: 都道府県が {len(out)} 件しか読めない")
    return out


def _parse_workbook(path: Path, fiscal_year: int) -> list[tuple]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: list[tuple] = []
    months = 0
    for sheet_name in workbook.sheetnames:
        year_month = check_sheet(sheet_name, fiscal_year)
        if year_month is None:
            continue
        year, month = year_month
        out.extend(
            _parse_sheet(list(workbook[sheet_name].iter_rows(values_only=True)),
                         sheet_name, year, month)
        )
        months += 1
    workbook.close()
    if not months:
        raise RuntimeError(f"{fiscal_year}年度のファイルに月次シートが無い")
    return out


def download_and_parse(csv_path: Path, work_dir: Path | None = None) -> int:
    """都道府県別発電所数、出力を取得・整形して CSV に書き出し、行数を返す。"""
    work_dir = Path(work_dir) if work_dir else csv_path.parent
    work_dir.mkdir(parents=True, exist_ok=True)

    all_rows: list[tuple] = []
    for fiscal_year, path in resolve_sources(TABLE_NO):
        xlsx_path = work_dir / f"power_plants_{fiscal_year}.xlsx"
        fetch_file(path, xlsx_path)
        rows = _parse_workbook(xlsx_path, fiscal_year)
        logger.info(f"  {fiscal_year}年度: {len(rows)} rows ({path})")
        all_rows.extend(rows)

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(OUTPUT_COLUMNS)
        writer.writerows(all_rows)

    return len(all_rows)

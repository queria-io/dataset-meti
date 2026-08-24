"""電力調査統計 電気事業者の発電所数、出力の取得・整形。

資源エネルギー庁が公表する統計表 1-(1)「電気事業者の発電所数、出力」の Excel を
年度ごとに取得し、事業者×月の 1 行 = 1 レコードへ整形して CSV に保存する。

都道府県別（1-(2)）と違い行が電気事業者で、水力の一般・揚水式と火力の燃料別 7 区分を
持つ。事業区分の列は発電実績（2-(1)）と同じ形で、電気事業法の改正で後から増えるため
列位置ではなく見出しで対応づける。見出しは 3 行で、3 行目の測定項目（発電所数・最大出力）が
値を持つ列にだけ入るので、事業区分の列との切り分けにも使う。
"""

import csv
import logging
import re
from pathlib import Path

import openpyxl

from enecho import fetch_file
from power_survey_common import (
    check_sheet,
    published_as_of,
    resolve_sources,
    squash,
)

logger = logging.getLogger("pipelines")

TABLE_NO = "1-1"

NAME_HEADING = "事業者名"
# 発電実績（2-(1)）と同じく、原典が「事業社名」と誤記している年度がある。
NAME_HEADINGS = {NAME_HEADING, "事業社名"}
TOTAL_ROW_HEADING = "合計"

# 事業区分の見出しと列名の対応。原典は該当すれば ○、しなければ 0 を置く。
# 配電事業者・特定卸供給事業者は 2022 年 4 月施行の電気事業法改正で新設された区分で、
# それ以前の年度には列そのものが無い。
FLAG_HEADINGS = {
    NAME_HEADING: "_name",
    "小売電気事業者": "is_retail",
    "一般送配電事業者": "is_general_transmission_distribution",
    "送電事業者": "is_transmission",
    "配電事業者": "is_distribution",
    "特定送配電事業者": "is_specified_transmission_distribution",
    "発電事業者": "is_generation",
    "特定卸供給事業者": "is_specified_wholesale",
}

# 大分類の表記ゆれ。火力は年度によって「火力発電」と「火力発電所」が混ざる。
TOP_ALIASES = {"火力発電": "火力発電所"}

# 発電所の種別（大分類, 小分類）と列名の前半の対応。小分類の〔〕は再掲を表す
# 原典の記法で、対応づけの前に外す。
CATEGORY_HEADINGS = {
    ("水力発電所", "一般"): "hydro_conventional",
    ("水力発電所", "揚水式"): "hydro_pumped_storage",
    ("水力発電所", "計"): "hydro",
    ("火力発電所", "石炭"): "thermal_coal",
    ("火力発電所", "ＬＮＧ"): "thermal_lng",
    ("火力発電所", "石油"): "thermal_oil",
    ("火力発電所", "ＬＰＧ"): "thermal_lpg",
    ("火力発電所", "その他ガス"): "thermal_other_gas",
    ("火力発電所", "歴青質混合物"): "thermal_bituminous",
    ("火力発電所", "その他"): "thermal_other",
    ("火力発電所", "計"): "thermal",
    ("原子力発電所", ""): "nuclear",
    ("新エネルギー等発電所", "風力"): "wind",
    ("新エネルギー等発電所", "太陽光"): "solar",
    ("新エネルギー等発電所", "地熱"): "geothermal",
    ("新エネルギー等発電所", "バイオマス"): "biomass",
    ("新エネルギー等発電所", "廃棄物"): "waste",
    ("新エネルギー等発電所", "蓄電池"): "storage_battery",
    ("新エネルギー等発電所", "計"): "new_energy",
    ("その他", ""): "other",
    ("計", ""): "total",
}

# 見出しの 3 行目。種別ごとに 2 列あり、値を持つ列にだけ入る。
MEASURE_HEADINGS = {"発電所数": "plants", "最大出力": "capacity_kw"}

FLAG_COLUMNS = [
    "is_retail",
    "is_general_transmission_distribution",
    "is_transmission",
    "is_distribution",
    "is_specified_transmission_distribution",
    "is_generation",
    "is_specified_wholesale",
]
VALUE_COLUMNS = [
    f"{category}_{suffix}"
    for category in CATEGORY_HEADINGS.values()
    for suffix in MEASURE_HEADINGS.values()
]

# 年度によって存在しない列。無ければ NULL で埋める（0 ではない）。
OPTIONAL_COLUMNS = {
    "is_distribution",
    "is_specified_wholesale",
    "storage_battery_plants",
    "storage_battery_capacity_kw",
}

OUTPUT_COLUMNS = [
    "year_month", "operator_name", *FLAG_COLUMNS, *VALUE_COLUMNS, "published_as_of",
]

# 事業区分の ○ は年度・行によって U+25CB と U+3007 が混ざる。該当しない欄は 0。
YES_MARKS = {"○", "〇"}

# 再掲のバイオマス・廃棄物は、数値ではなく〔〕や［］で括った文字列で入っている月がある。
# 括弧と桁区切りを外した中身が数値なら値として読む（〔　　〕のような空欄は欠測）。
VALUE_TEXT_PATTERN = re.compile(r"^[〔［\[]?([\d,]+(?:\.\d+)?)[〕］\]]?$")


def _value(cell: object) -> float | None:
    """値のセルを数値にする。読めないセル（空欄・記号のみ）は欠測。"""
    if isinstance(cell, (int, float)):
        return float(cell)
    if isinstance(cell, str):
        matched = VALUE_TEXT_PATTERN.match(squash(cell))
        if matched:
            return float(matched.group(1).replace(",", ""))
    return None


def _strip_restatement(heading: str) -> str:
    """再掲を表す括弧を外す。〔バイオマス〕は年度によって括弧の有無が変わる。"""
    return heading.strip("〔〕［］[]")


def _header_row(rows: list[tuple], sheet_name: str) -> tuple[int, int]:
    """見出し行の位置と事業者名が入る列を返す。"""
    for index, row in enumerate(rows):
        for column, value in enumerate(row):
            if squash(value) in NAME_HEADINGS:
                return index, column
    raise RuntimeError(f"{sheet_name}: 見出し行が見つからない")


def _resolve_columns(rows: list[tuple], sheet_name: str) -> tuple[int, dict[str, int]]:
    """見出しから列名と列位置の対応を作る。

    大分類・小分類は結合セルなので右へ引き延ばす。測定項目の見出しがある列は値の列で、
    （大分類, 小分類）と測定項目から列名を組み立てる。測定項目の無い列のうち大分類が
    自分の列に置かれているものが事業区分で、それ以外は表の外側なので捨てる。
    """
    header, name_column = _header_row(rows, sheet_name)
    if len(rows) < header + 3:
        raise RuntimeError(f"{sheet_name}: 見出しが 3 行に足りない")
    top_row, sub_row, measure_row = rows[header], rows[header + 1], rows[header + 2]

    resolved: dict[str, int] = {}
    top = ""
    sub = ""
    width = max(len(top_row), len(sub_row), len(measure_row))
    for column in range(width):
        own_top = squash(top_row[column] if column < len(top_row) else None)
        if own_top in NAME_HEADINGS:
            own_top = NAME_HEADING
        own_top = TOP_ALIASES.get(own_top, own_top)
        if own_top:
            top, sub = own_top, ""
        sub = _strip_restatement(
            squash(sub_row[column] if column < len(sub_row) else None)
        ) or sub
        suffix = MEASURE_HEADINGS.get(squash(measure_row[column])
                                      if column < len(measure_row) else "")
        if suffix is None:
            if not own_top:
                continue
            name = FLAG_HEADINGS.get(own_top)
            if name is None:
                raise RuntimeError(f"{sheet_name}: 未知の見出し '{own_top}'（列 {column}）")
        else:
            category = CATEGORY_HEADINGS.get((top, sub))
            if category is None:
                raise RuntimeError(
                    f"{sheet_name}: 未知の見出し '{top}' / '{sub}'（列 {column}）"
                )
            name = f"{category}_{suffix}"
        if name in resolved:
            raise RuntimeError(f"{sheet_name}: 見出し '{name}' が重複している（列 {column}）")
        resolved[name] = column

    expected = set(FLAG_HEADINGS.values()) | set(VALUE_COLUMNS)
    missing = expected - OPTIONAL_COLUMNS - set(resolved)
    if missing:
        raise RuntimeError(f"{sheet_name}: 見出しが足りない（{'・'.join(sorted(missing))}）")
    if resolved["_name"] != name_column:
        raise RuntimeError(f"{sheet_name}: 事業者名の列が定まらない")
    return header, resolved


def _operator_name(cell: object) -> str | None:
    """事業者名のセルを名前にする。空欄は行そのものが無いものとして None。

    原典は事業者名の欄に法人番号（数値）を書いている行がある（2026年4月に1件）。
    文字列だけを拾うと値のある行が黙って落ちるので、数値もそのまま名前として扱う。
    """
    if isinstance(cell, str):
        return cell.strip() or None
    if isinstance(cell, int):
        return str(cell)
    if isinstance(cell, float):
        return str(int(cell)) if cell.is_integer() else str(cell)
    return None


def _flag(cell: object) -> bool:
    """事業区分の欄。○ が該当で、0 や空欄は非該当。"""
    return isinstance(cell, str) and cell.strip() in YES_MARKS


def _parse_sheet(rows: list[tuple], sheet_name: str, year: int, month: int) -> list[tuple]:
    """1 か月のシートを事業者ごとの行リストへ展開する。

    同じ事業者名が 2 行並ぶ月がある（原典どおり）ので、事業者名は月内でも一意にならない。
    """
    header, columns = _resolve_columns(rows, sheet_name)
    name_column = columns.pop("_name")
    as_of = published_as_of(rows[0]) if rows else None
    year_month = f"{year}{month:02d}"
    mapped = set(columns.values()) | {name_column}

    out: list[tuple] = []
    for row in rows[header + 3:]:
        name = _operator_name(row[name_column] if name_column < len(row) else None)
        if name is None:
            continue
        # 原典は事業者の並びの末尾に合計行を置き、その下に備考が続く。
        if squash(name) == TOTAL_ROW_HEADING:
            break
        # 見出しの無い列に値があれば読み落としているので気付けるようにする。
        for column, cell in enumerate(row):
            if column not in mapped and isinstance(cell, (int, float)) and cell:
                raise RuntimeError(f"{sheet_name}: 見出しの無い列 {column} に値がある")
        flags = []
        for column in FLAG_COLUMNS:
            index = columns.get(column)
            # その年度に区分そのものが無ければ「該当しない」ではなく不明。
            flags.append(None if index is None
                         else _flag(row[index] if index < len(row) else None))
        values = []
        for column in VALUE_COLUMNS:
            index = columns.get(column)
            cell = row[index] if index is not None and index < len(row) else None
            values.append(_value(cell))
        out.append((year_month, name.strip(), *flags, *values, as_of))

    if not out:
        raise RuntimeError(f"{sheet_name}: 事業者の行が読めない")
    return out


def _parse_workbook(path: Path, fiscal_year: int) -> list[tuple]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: list[tuple] = []
    months = 0
    for sheet_name in workbook.sheetnames:
        year_month = check_sheet(sheet_name, fiscal_year)
        if year_month is None:
            continue
        year, month = year_month
        out.extend(
            _parse_sheet(list(workbook[sheet_name].iter_rows(values_only=True)),
                         sheet_name, year, month)
        )
        months += 1
    workbook.close()
    if not months:
        raise RuntimeError(f"{fiscal_year}年度のファイルに月次シートが無い")
    return out


def download_and_parse(csv_path: Path, work_dir: Path | None = None) -> int:
    """電気事業者の発電所数、出力を取得・整形して CSV に書き出し、行数を返す。"""
    work_dir = Path(work_dir) if work_dir else csv_path.parent
    work_dir.mkdir(parents=True, exist_ok=True)

    all_rows: list[tuple] = []
    for fiscal_year, path in resolve_sources(TABLE_NO):
        xlsx_path = work_dir / f"power_plants_operator_{fiscal_year}.xlsx"
        fetch_file(path, xlsx_path)
        rows = _parse_workbook(xlsx_path, fiscal_year)
        logger.info(f"  {fiscal_year}年度: {len(rows)} rows ({path})")
        all_rows.extend(rows)

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(OUTPUT_COLUMNS)
        writer.writerows(all_rows)

    return len(all_rows)
